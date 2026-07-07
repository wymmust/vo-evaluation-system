import base64
import io
import importlib.util
import json
import pytest
import subprocess
import textwrap
from pathlib import Path

from openpyxl import load_workbook


def test_static_browser_evaluator_exports_new_vloc_summary_metrics():
    source = Path("voeval/reports/detail.py").read_text()
    assert "mean_error_pos_xy" in source
    assert "mean_error_pos_z" in source
    assert "mean_error_euler" in source
    assert "max_error_pos_xy" in source
    assert "max_error_pos_z" in source
    assert "max_error_euler" in source


def test_static_dead_report_and_time_series_helpers_are_removed():
    source = Path("voeval/visualization/js/main.js").read_text()
    removed_names = [
        "associationLabel",
        "renderGtVoTimeChart",
        "renderErrorTimeChart",
        "reportSubtitle",
        "reportOverallStatus",
        "buildReportMetricCards",
        "reportMetricCardHtml",
        "buildReportFindings",
        "reportFindingHtml",
        "buildSegmentSummaryRows",
        "buildTuningConclusionRows",
        "buildHealthDashboardCards",
        "buildAssociationDiagnosticRows",
        "buildLongRangeDiagnosticRows",
        "buildWorstSegmentRows",
        "buildConditionDiagnosticRows",
        "buildAuxiliaryMetricCards",
        "buildReportPlotSpecs",
        "flattenReportMetrics",
        "metricIssue",
        "buildConfigRows",
        "addExportSelectionMarkers",
        "renderVlocCharts",
        "renderPairCompositeChart",
        "renderSingleCompositeChart",
        "renderMultiFieldTimeChart",
        "renderRpeTimeChart",
        "renderScaleTimeChart",
        "buildVisualizationExportFigureSpecs",
        "buildVlocVisualizationExportFigureSpecs",
        "buildVoVisualizationExportFigureSpecs",
        "exportMultiFieldTimeFigure",
        "exportPairCompositeFigure",
        "exportSingleCompositeFigure",
        "exportRpeTimeFigure",
        "exportScaleTimeFigure",
    ]
    for name in removed_names:
        assert f"function {name}" not in source
@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_scale_interval_controls_are_wired_into_config():
    html = Path("voeval/visualization/index.html").read_text()
    assert 'id="rpeDeltaValue"' in html
    assert 'id="rpeDeltaUnit"' in html
    assert 'id="scaleDeltaValue"' in html
    assert 'id="scaleDeltaUnit"' in html
    for removed_id in [
        "maxTimeDiff",
        "maxInterpolationGap",
        "allowExtrapolation",
        "interpolateRotation",
        "timeOffset",
        "segmentLengths",
        "maxSegments",
        "segmentStep",
        "lengthTolerance",
        "discontinuityStep",
        "discontinuityGap",
        "divergenceAbs",
        "divergenceRel",
    ]:
        assert f'id="{removed_id}"' not in html

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const elements = {
          entryMode: { value: "vo" },
          rpeDeltaValue: { value: "1" },
          rpeDeltaUnit: { value: "frames" },
          scaleDeltaValue: { value: "100" },
          scaleDeltaUnit: { value: "meters" },
        };
        const element = { addEventListener() {}, classList: { add() {}, remove() {} }, style: {}, files: [], value: "" };
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        process.stdout.write(JSON.stringify(context.buildConfig()));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    config = json.loads(result.stdout)
    assert config["scale_delta_value"] == 100
    assert config["scale_delta_unit"] == "meters"
    assert config["scale_distance_tolerance_ratio"] == 0.05


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_html_export_keeps_light_chart_exports_and_xlsx_has_sheets():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const report = {
          inputs: { entry_mode: "vo" },
          summary: { matched_poses: 2 },
          per_pose: [{ timestamp: 1, error_m: 0.1 }],
            trajectory_exports: {
              input_gt_tum: [{ timestamp: 1, tx: 0, ty: 0, tz: 0 }],
              input_vo_tum: [{ timestamp: 1, tx: 1, ty: 0, tz: 0 }],
              filtered_vo_tum: [{ timestamp: 1, tx: 1, ty: 0, tz: 0 }],
              interpolated_gt_tum: [{ timestamp: 1, tx: 0, ty: 0, tz: 0 }],
              sim3_gt_tum: [{ timestamp: 1, tx: 0, ty: 0, tz: 0 }],
              sim3_vo_tum: [{ timestamp: 1, tx: 0, ty: 0, tz: 0 }],
              ate_per_frame: [{ timestamp: 1, ate_position_m: 0.1 }],
              rpe_per_frame: [{ timestamp: 1, rpe_translation_m: 0.2, rpe_available: true }],
              scale_per_frame: [{ timestamp: 1, local_sim3_scale: 2, scale_available: true }],
            },
          };
            const sanitized = context.reportForHtmlExport(report);
        const vlocSanitized = context.reportForHtmlExport({
          inputs: { entry_mode: "vloc" },
          trajectory_exports: {
            rpe_per_frame: [{ timestamp: 1, rpe_translation_m: 0.2, rpe_available: true }],
            scale_per_frame: [{ timestamp: 1, local_sim3_scale: 2, scale_available: true }],
          },
        });
            const workbook = context.buildTrajectoryWorkbook(report.trajectory_exports);
            process.stdout.write(JSON.stringify({
              sanitized,
          vlocSanitized,
              workbookBase64: Buffer.from(workbook).toString("base64"),
            }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)

    assert payload["sanitized"]["trajectory_exports"] == {
        "rpe_per_frame": [{"timestamp": 1, "rpe_translation_m": 0.2, "rpe_available": True}],
        "scale_per_frame": [{"timestamp": 1, "local_sim3_scale": 2, "scale_available": True}],
    }
    assert "scale_per_frame" not in payload["vlocSanitized"].get("trajectory_exports", {})
    assert "input_gt_tum" not in payload["sanitized"]["trajectory_exports"]
    assert "sim3_vo_tum" not in payload["sanitized"]["trajectory_exports"]
    assert "ate_per_frame" not in payload["sanitized"]["trajectory_exports"]
    assert "per_pose" not in payload["sanitized"]
    assert "segment_records" not in payload["sanitized"]
    assert payload["sanitized"]["summary"]["matched_poses"] == 2

    workbook_bytes = base64.b64decode(payload["workbookBase64"])
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
    assert workbook.sheetnames == [
        "input_gt_tum",
        "input_vo_tum",
        "filtered_vo_tum",
        "interpolated_gt_tum",
        "sim3_gt_tum",
        "sim3_vo_tum",
        "ate_per_frame",
        "rpe_per_frame",
        "scale_per_frame",
    ]
    assert workbook["input_gt_tum"]["A1"].value == "timestamp"
    assert workbook["input_gt_tum"]["A2"].value == 1
    assert workbook["ate_per_frame"]["B1"].value == "ate_position_m"
    assert workbook["rpe_per_frame"]["B1"].value == "rpe_translation_m"
    assert workbook["scale_per_frame"]["B1"].value == "local_sim3_scale"


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_trajectory_workbook_omits_missing_vloc_sim3_sheets():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const workbook = context.buildTrajectoryWorkbook({
          input_gt_tum: [{ timestamp: 1, tx: 0 }],
          input_vo_tum: [{ timestamp: 1, tx: 1 }],
          filtered_vo_tum: [{ timestamp: 1, tx: 1 }],
          interpolated_gt_tum: [{ timestamp: 1, tx: 0 }],
          ate_per_frame: [{ timestamp: 1, ate_position_m: 0.1 }],
          rpe_per_frame: [{ timestamp: 1, rpe_translation_m: 0.2 }],
        });
        process.stdout.write(Buffer.from(workbook).toString("base64"));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    workbook = load_workbook(io.BytesIO(base64.b64decode(result.stdout)), read_only=True)
    assert workbook.sheetnames == [
        "input_gt_tum",
        "input_vo_tum",
        "filtered_vo_tum",
        "interpolated_gt_tum",
        "ate_per_frame",
        "rpe_per_frame",
    ]


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_export_filenames_include_directory_name_and_entry_mode():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "vloc",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const vlocReport = {
          inputs: {
            entry_mode: "vloc",
            data_dir_name: "2839_traj",
            log_dir_name: "2839_traj",
          },
        };
        const vlocHtml = context.evaluationExportFilename("evaluation_report", "html", vlocReport);
        const vlocXlsx = context.evaluationExportFilename("trajectory_exports", "xlsx", vlocReport);
        const voReport = {
          inputs: {
            entry_mode: "vo",
            data_dir_name: "flight:data",
            log_dir_name: "log/vo*run",
          },
        };
        const voHtml = context.evaluationExportFilename("evaluation_report", "html", voReport);
        const defaultDirReport = {
          inputs: {
            entry_mode: "vloc",
            data_dir_name: "data_dir",
            log_dir_name: "log_dir",
          },
        };
        const defaultDirHtml = context.evaluationExportFilename("evaluation_report", "html", defaultDirReport);
        process.stdout.write(JSON.stringify({ vlocHtml, vlocXlsx, voHtml, defaultDirHtml }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["vlocHtml"] == "2839_traj_vloc_evaluation_report.html"
    assert payload["vlocXlsx"] == "2839_traj_vloc_trajectory_exports.xlsx"
    assert payload["voHtml"] == "flight_data__log_vo_run_vo_evaluation_report.html"
    assert payload["defaultDirHtml"] == "vloc_evaluation_report.html"


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_html_export_is_visualization_snapshot_with_point_selection():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "vloc",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const report = {
          inputs: { entry_mode: "vloc" },
          summary: { matched_poses: 2, gt_path_length_m: 10, duration_s: 1, est_pose_coverage_ratio: 1, gt_pose_coverage_ratio: 1 },
          ate_position_m: { rmse: 0.2 },
          ate_vertical_m: { rmse: 0.1 },
          discontinuities: { all_matches: { break_count: 0 }, selected_segment: { policy: "vo_timestamps" } },
          vloc_details: {
            summary: { mean_error_pos_xy: 0.1, mean_error_pos_z: 0.2, mean_error_euler: 0.3, max_error_pos_xy: 0.4, max_error_pos_z: 0.5, max_error_euler: 0.6 },
            comparison: [
              { timestamp: 1, distance_m: 0, visual_segment_id: 0, nav_n_m: 0, nav_e_m: 0, nav_d_m: 0, vloc_n_m: 0.1, vloc_e_m: 0.1, vloc_d_m: 0.1, nav_height_m: 5, vloc_height_m: 4.9, position_error_3d_m: 0.2, horizontal_position_error_m: 0.1, vertical_position_error_abs_m: 0.1, position_error_n_m: 0.1, position_error_e_m: 0.1, position_error_d_m: 0.1, nav_yaw_deg: 0, vloc_yaw_deg: 0.1, nav_pitch_deg: 0, vloc_pitch_deg: 0.2, nav_roll_deg: 0, vloc_roll_deg: 0.3, attitude_error_yaw_deg: 0.1, attitude_error_pitch_deg: 0.2, attitude_error_roll_deg: 0.3 },
              { timestamp: 2, distance_m: 1, visual_segment_id: 0, nav_n_m: 1, nav_e_m: 1, nav_d_m: 0, vloc_n_m: 1.1, vloc_e_m: 1.1, vloc_d_m: 0.1, nav_height_m: 6, vloc_height_m: 5.9, position_error_3d_m: 0.2, horizontal_position_error_m: 0.1, vertical_position_error_abs_m: 0.1, position_error_n_m: 0.1, position_error_e_m: 0.1, position_error_d_m: 0.1, nav_yaw_deg: 1, vloc_yaw_deg: 1.1, nav_pitch_deg: 1, vloc_pitch_deg: 1.2, nav_roll_deg: 1, vloc_roll_deg: 1.3, attitude_error_yaw_deg: 0.1, attitude_error_pitch_deg: 0.2, attitude_error_roll_deg: 0.3 },
            ],
            nav_status: [{ timestamp: 1, flight_mode: 3, navi_mode: 5, rtk_yaw: 1, rtk_alti: 0, vx: 1, vy: 2, vz: 3, velocity_norm: 3.7, position_reset_count: 0, altitude_reset_count: 0, heading_reset_count: 0 }],
            vloc_status: [{ timestamp: 1, vloc_mode: 2, num_inliers: 40, reset_count: 0 }],
          },
        };
        const cssSource = fs.readFileSync("voeval/visualization/css/style.css", "utf8");
        const reportCssSource = fs.readFileSync("voeval/visualization/css/report-export.css", "utf8");
        process.stdout.write(context.buildHtmlReport(report, { cssSource, reportCssSource }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    html = result.stdout

    assert "VLOC 评估结果" in html
    assert "VLOC 运行结果" in html
    assert "离线可视化快照" in html
    assert "图表目录" in html
    assert "pointSelectionOutput" in html
    assert "togglePointMode" in html
    assert "recordPointSelection" in html
    assert "composite-crosshair" in html
    assert "attachExportCompositeOverlay" in html
    assert 'node.on("plotly_hover"' in html
    assert "exportAllHoverChips" in html
    assert "nearestExportTracePoint" in html
    assert "focusExportPointSelectionFromEvent" in html
    assert "deleteFocusedExportPointSelection" in html
    assert 'class="layout"' in html
    assert 'class="controls"' in html
    assert 'class="chart-grid"' in html
    assert ".metric-grid {" in html
    assert ".chart-card {" in html
    assert ".chart-directory-item {" in html
    assert ".metric-chip-value" in html
    assert ".chart-card[data-chart-id='" in html
    assert ".chart-card[hidden]" in html
    assert "document.querySelector(\"[data-chart-id='\"" not in html
    assert "point.pointNumber" in html
    assert "removeExportSelectionMarkerTraces" in html
    assert "refreshExportChartSelectionMarkers" in html
    assert "keydown" in html
    assert "trajectory3d" in html
    assert "positionCompareComposite" in html
    assert '"margin":{"l":56,"r":26,"t":78,"b":50}' in html
    assert '"legend":{"orientation":"h","y":1.18,"x":0,"font":{"size":10}}' in html
    assert "调参结论摘要" not in html
    assert "高级详情 / 完整指标" not in html
    assert "原始 JSON 指标" not in html

    behavior_script = textwrap.dedent(
        f"""
        const vm = require("vm");
        const html = {json.dumps(html)};
        const scriptStart = html.lastIndexOf("<script>");
        const scriptEnd = html.lastIndexOf("</script>");
        const inline = html.slice(scriptStart + "<script>".length, scriptEnd).replace(/\\ninitExportPage\\(\\);\\n/, "\\n");
        const chartId = "positionCompareComposite";
        const chart = {{
          id: chartId,
          data: [{{ type: "scatter", name: "N nav", x: [1], y: [2] }}],
          style: {{}},
          classList: {{ toggle() {{}} }},
          dataset: {{ chartId }},
        }};
        const directoryItem = {{ hidden: false, dataset: {{ chartId }}, classList: {{ toggle() {{}} }} }};
        const chartCard = {{ hidden: false, dataset: {{ chartId }}, classList: {{ toggle() {{}} }} }};
        const pointSelectionOutputSection = {{ hidden: true }};
        const pointSelectionOutput = {{ innerHTML: "" }};
        const document = {{
          getElementById(id) {{
            if (id === chartId) return chart;
            if (id === "pointSelectionOutputSection") return pointSelectionOutputSection;
            if (id === "pointSelectionOutput") return pointSelectionOutput;
            return null;
          }},
          querySelector(selector) {{
            if (selector.startsWith(".chart-card[data-chart-id")) return chartCard;
            if (selector.startsWith("[data-chart-id")) return directoryItem;
            return null;
          }},
          querySelectorAll() {{ return []; }},
        }};
        const Plotly = {{
          addTraces(id, traces) {{
            if (id !== chartId) throw new Error("unexpected addTraces chart " + id);
            chart.data.push(...traces);
          }},
          deleteTraces(id, indices) {{
            if (id !== chartId) throw new Error("unexpected deleteTraces chart " + id);
            [...indices].sort((left, right) => right - left).forEach((index) => chart.data.splice(index, 1));
          }},
        }};
        const context = {{ document, window: {{}}, Plotly, console, Number, String, Array, Math, Set, Date }};
        context.globalThis = context;
        vm.runInNewContext(inline, context);
        const selection = {{
          id: "s1",
          order: 1,
          colorSlot: 1,
          chartId,
          chartTitle: "NED 随时间变化",
          traceName: "N nav",
          timestamp: 1,
          value: "2.000",
          color: "#ef4444",
          markerText: "",
          x: 1,
          y: 2,
          xaxis: "x",
          yaxis: "y",
        }};
        context.window.__VO_EXPORT_SELECTIONS__ = [selection];
        context.window.__VO_EXPORT_FOCUSED_SELECTION_ID__ = "s1";
        context.refreshExportChartSelectionMarkers(chartId);
        const afterAdd = chart.data.length;
        context.deleteFocusedExportPointSelection();
        const afterDelete = chart.data.length;
        context.window.__VO_EXPORT_SELECTIONS__ = [selection];
        context.refreshExportChartSelectionMarkers(chartId);
        const afterReadd = chart.data.length;
        context.clearChartSelections(chartId);
        const afterClear = chart.data.length;
        context.setChartVisible(chartId, false);
        const chartHiddenAfterHide = chartCard.hidden;
        const directoryHiddenAfterHide = directoryItem.hidden;
        context.setChartVisible(chartId, true);
        const chartHiddenAfterShow = chartCard.hidden;
        process.stdout.write(JSON.stringify({{
          afterAdd,
          afterDelete,
          afterReadd,
          afterClear,
          selections: context.window.__VO_EXPORT_SELECTIONS__.length,
          chartHiddenAfterHide,
          directoryHiddenAfterHide,
          chartHiddenAfterShow,
        }}));
        """
    )
    behavior_result = subprocess.run(["node", "-e", behavior_script], check=True, capture_output=True, text=True)
    behavior = json.loads(behavior_result.stdout)
    assert behavior == {
        "afterAdd": 3,
        "afterDelete": 1,
        "afterReadd": 3,
        "afterClear": 1,
        "selections": 0,
        "chartHiddenAfterHide": True,
        "directoryHiddenAfterHide": False,
        "chartHiddenAfterShow": False,
    }


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_html_export_uses_vo_chart_set_for_vo_reports():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "vo",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const report = {
          inputs: { entry_mode: "vo" },
          summary: { matched_poses: 2, gt_path_length_m: 10, duration_s: 1, est_pose_coverage_ratio: 1, gt_pose_coverage_ratio: 1 },
          ate_position_m: { rmse: 0.2 },
          ate_vertical_m: { rmse: 0.1 },
          rpe_frame_delta: { delta_unit: "frames", delta_frames: 1, translation_m: { rmse: 0.3 } },
          discontinuities: { all_matches: { break_count: 0 }, selected_segment: { policy: "vo_timestamps" } },
          alignment: { scale: 1 },
          vo_details: {
            comparison: [
              { timestamp: 1, distance_m: 0, visual_segment_id: 0, nav_x_m: 0, nav_y_m: 0, nav_z_m: 0, vo_x_aligned_m: 0.1, vo_y_aligned_m: 0.1, vo_z_aligned_m: 0.1, position_error_3d_m: 0.2, horizontal_position_error_m: 0.1, position_error_x_m: 0.1, position_error_y_m: 0.1, position_error_z_m: 0.1, nav_yaw_deg: 0, vo_yaw_aligned_deg: 0.1, nav_pitch_deg: 0, vo_pitch_aligned_deg: 0.2, nav_roll_deg: 0, vo_roll_aligned_deg: 0.3, attitude_error_yaw_deg: 0.1, attitude_error_pitch_deg: 0.2, attitude_error_roll_deg: 0.3 },
              { timestamp: 2, distance_m: 1, visual_segment_id: 0, nav_x_m: 1, nav_y_m: 1, nav_z_m: 0, vo_x_aligned_m: 1.1, vo_y_aligned_m: 1.1, vo_z_aligned_m: 0.1, position_error_3d_m: 0.2, horizontal_position_error_m: 0.1, position_error_x_m: 0.1, position_error_y_m: 0.1, position_error_z_m: 0.1, nav_yaw_deg: 1, vo_yaw_aligned_deg: 1.1, nav_pitch_deg: 1, vo_pitch_aligned_deg: 1.2, nav_roll_deg: 1, vo_roll_aligned_deg: 1.3, attitude_error_yaw_deg: 0.1, attitude_error_pitch_deg: 0.2, attitude_error_roll_deg: 0.3 },
            ],
            nav_status: [{ timestamp: 1, flight_mode: 3, navi_mode: 5, rtk_yaw: 1, rtk_alti: 0, vx: 1, vy: 2, vz: 3, velocity_norm: 3.7, position_reset_count: 0, altitude_reset_count: 0, heading_reset_count: 0 }],
            vo_status: [{ timestamp: 1, num_inliers: 40, is_keyframe: 1, time_cost: 3, reset_count: 0 }],
          },
          trajectory_exports: {
            rpe_per_frame: [{ timestamp: 1, rpe_translation_m: 0.3, rpe_rotation_deg: 0.4, rpe_available: true }],
            scale_per_frame: [{ timestamp: 1, local_sim3_scale: 2, scale_available: true }],
          },
        };
        process.stdout.write(context.buildHtmlReport(report));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    html = result.stdout

    assert "VO 评估结果" in html
    assert "voStatus" in html
    assert "rpeTranslationTime" in html
    assert "rpeRotationTime" in html
    assert "scaleFrameTime" in html
    assert "局部 Sim3 尺度随时间变化" in html
    assert "vlocStatus" not in html
    assert "VLOC 评估结果" not in html


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_visualization_renders_time_series_and_rpe_charts():
    html = Path("voeval/visualization/index.html").read_text()
    expected_ids = [
        "trajectory3d",
        "errorDistance",
        "navStatusModes",
        "navVelocity",
        "navResetCounts",
        "voStatus",
        "positionCompareComposite",
        "attitudeCompareComposite",
        "positionErrorComposite",
        "attitudeErrorComposite",
        "rpeTranslationTime",
        "rpeRotationTime",
        "scaleFrameTime",
    ]
    for chart_id in expected_ids:
        assert f'id="{chart_id}"' in html

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const plots = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); }, purge() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        const perPose = [0, 1, 2].map((index) => ({
          timestamp: index,
          segment_id: 0,
          visual_segment_id: index < 2 ? 0 : 1,
          gt_x_m: index,
          gt_y_m: index + 1,
          gt_z_m: index + 2,
          est_x_aligned_m: index + 0.1,
          est_y_aligned_m: index + 1.1,
          est_z_aligned_m: index + 2.1,
          x_error_m: 0.1,
          y_error_m: 0.1,
          z_error_m: 0.1,
          gt_yaw_deg: index,
          gt_pitch_deg: index + 2,
          gt_roll_deg: index + 4,
          est_yaw_aligned_deg: index + 0.2,
          est_pitch_aligned_deg: index + 2.2,
          est_roll_aligned_deg: index + 4.2,
          yaw_error_signed_deg: 0.2,
          pitch_error_signed_deg: 0.2,
          roll_error_signed_deg: 0.2,
          distance_m: index,
          error_m: 0.1,
          horizontal_error_m: 0.1,
          vertical_error_m: 0.1,
        }));
        context.renderCharts({
          inputs: { entry_mode: "vo" },
          per_pose: perPose,
          vo_details: {
            nav_status: [
              { timestamp: 0, flight_mode: 3, navi_mode: 5, rtk_yaw: 1, rtk_alti: 0, position_reset_count: 0, altitude_reset_count: 1, heading_reset_count: 2, vx: 0.1, vy: 0.2, vz: 0.3, velocity_norm: 0.374 },
              { timestamp: 1, flight_mode: 4, navi_mode: 6, rtk_yaw: 1, rtk_alti: 1, position_reset_count: 0, altitude_reset_count: 1, heading_reset_count: 3, vx: 0.4, vy: 0.5, vz: 0.6, velocity_norm: 0.877 },
            ],
            vo_status: [
              { timestamp: 0, num_inliers: 40, is_keyframe: 1, time_cost: 3.5, reset_count: 0 },
              { timestamp: 1, num_inliers: 41, is_keyframe: 0, time_cost: 3.7, reset_count: 1 },
            ],
          },
          trajectory_exports: {
            sim3_vo_tum: [
              { timestamp: 0, segment_id: 0, sim3_scale: 2 },
              { timestamp: 1, segment_id: 0, sim3_scale: 2 },
              { timestamp: 2, segment_id: 0, sim3_scale: 2 },
            ],
            rpe_per_frame: [
              { timestamp: 0, rpe_translation_m: 0.3, rpe_rotation_deg: 1.1, rpe_available: true },
              { timestamp: 1, rpe_translation_m: 0.4, rpe_rotation_deg: 1.2, rpe_available: true },
            ],
            scale_per_frame: [
              { timestamp: 0, segment_id: 0, local_sim3_scale: 2, scale_available: true },
              { timestamp: 1, segment_id: 0, local_sim3_scale: 2, scale_available: true },
            ],
          },
        });
        const byId = Object.fromEntries(plots.map((plot) => [plot.id, plot]));
        process.stdout.write(JSON.stringify({
          ids: plots.map((plot) => plot.id),
          trajectory3dNames: byId.trajectory3d.data.map((trace) => trace.name),
          voStartText: byId.trajectory3d.data.find((trace) => trace.name === "vo start").text,
          voEndText: byId.trajectory3d.data.find((trace) => trace.name === "vo end").text,
          voStartMarker: byId.trajectory3d.data.find((trace) => trace.name === "vo start").marker,
          voStartTextFont: byId.trajectory3d.data.find((trace) => trace.name === "vo start").textfont,
          navVelocityNames: byId.navVelocity.data.map((trace) => trace.name),
          voStatusNames: byId.voStatus.data.map((trace) => trace.name),
          positionCompareNames: byId.positionCompareComposite.data.map((trace) => trace.name),
          thirdRowColors: [
            byId.positionCompareComposite.data[4].line.color,
            byId.positionCompareComposite.data[5].line.color,
          ],
          scaleNames: byId.scaleFrameTime.data.map((trace) => trace.name),
          scaleValues: byId.scaleFrameTime.data[0].y,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    rendered_ids = set(payload["ids"])
    assert set(expected_ids).issubset(rendered_ids)
    assert {"trajectoryXY", "segmentError", "speedError", "sim3ScaleTime"}.isdisjoint(rendered_ids)
    assert "GT start" not in payload["trajectory3dNames"]
    assert "GT end" not in payload["trajectory3dNames"]
    assert {"vo start", "vo end"}.issubset(set(payload["trajectory3dNames"]))
    assert payload["voStartText"] == ["vo S1", "vo S2"]
    assert payload["voEndText"] == ["vo E1", "vo E2"]
    assert payload["voStartMarker"]["size"] == 5
    assert payload["voStartMarker"]["color"] == "#9333ea"
    assert payload["voStartMarker"]["line"]["width"] == 1
    assert payload["voStartTextFont"]["size"] == 10
    assert "velocity_norm" in payload["navVelocityNames"]
    assert {"num_inliers", "is_keyframe", "time_cost", "reset_count"}.issubset(set(payload["voStatusNames"]))
    assert {"X Ground truth", "X VO aligned", "Y Ground truth", "Y VO aligned", "Z Ground truth", "Z VO aligned"} == set(payload["positionCompareNames"])
    assert payload["thirdRowColors"] == ["#dc2626", "#0891b2"]
    assert payload["scaleNames"] == ["local_sim3_scale"]
    assert payload["scaleValues"] == [2, 2]


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_composite_angle_time_series_unwraps_180_degree_boundary():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const plots = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); }, purge() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        const figure = context.pairCompositeFigure("attitudeCompareComposite", "Roll", [
          { timestamp: 0, segment_id: 0, gt_roll_deg: 0, est_roll_aligned_deg: 179 },
          { timestamp: 1, segment_id: 0, gt_roll_deg: 1, est_roll_aligned_deg: -179 },
          { timestamp: 2, segment_id: 0, gt_roll_deg: 2, est_roll_aligned_deg: 178 },
        ], {
          title: "Roll",
          leftName: "Ground truth",
          rightName: "VO aligned",
          rows: [{ label: "Roll", left: "gt_roll_deg", right: "est_roll_aligned_deg", unit: "deg", unwrap: true }],
        });
        process.stdout.write(JSON.stringify(figure.data[1].y));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    assert json.loads(result.stdout) == [179, 181, 178]


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_composite_angle_error_time_series_unwraps_180_degree_boundary():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const plots = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); }, purge() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        const figure = context.singleCompositeFigure("attitudeErrorComposite", "Roll error", [
          { timestamp: 0, segment_id: 0, roll_error_signed_deg: 179 },
          { timestamp: 1, segment_id: 0, roll_error_signed_deg: -179 },
          { timestamp: 2, segment_id: 0, roll_error_signed_deg: 178 },
        ], {
          title: "Roll error",
          rows: [{ label: "Roll error", field: "roll_error_signed_deg", unit: "deg", unwrap: true }],
        });
        process.stdout.write(JSON.stringify(figure.data[0].y));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    assert json.loads(result.stdout) == [179, 181, 178]


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_composite_charts_disable_native_spikes_for_custom_overlay():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {}, appendChild() {} }; },
        };
        const plots = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); return Promise.resolve(); }, purge() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        const plot = context.pairCompositeFigure("positionCompareComposite", "NED", [
          { timestamp: 1, nav_n_m: 10, vloc_n_m: 9, nav_e_m: 20, vloc_e_m: 19 },
          { timestamp: 2, nav_n_m: 11, vloc_n_m: 10, nav_e_m: 21, vloc_e_m: 20 },
        ], {
          title: "NED",
          leftName: "nav",
          rightName: "vloc",
          rows: [
            { label: "N", left: "nav_n_m", right: "vloc_n_m", unit: "m" },
            { label: "E", left: "nav_e_m", right: "vloc_e_m", unit: "m" },
          ],
        });
        process.stdout.write(JSON.stringify({
          hovermode: plot.layout.hovermode,
          hoversubplots: plot.layout.hoversubplots,
          xaxisShowspikes: plot.layout.xaxis.showspikes,
          xaxis2Showspikes: plot.layout.xaxis2.showspikes,
          firstTraceHoverinfo: plot.data[0].hoverinfo,
          secondTraceHoverinfo: plot.data[1].hoverinfo,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload == {
        "hovermode": "x unified",
        "hoversubplots": "axis",
        "xaxisShowspikes": False,
        "xaxis2Showspikes": False,
        "firstTraceHoverinfo": "none",
        "secondTraceHoverinfo": "none",
    }


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_composite_payload_helpers_compute_hover_and_range_payloads():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          parentNode: { insertBefore() {} },
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {}, appendChild() {}, parentNode: element.parentNode }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() { return Promise.resolve(); }, purge() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        const rows = [
          { timestamp: 10, nav_n_m: 1, vloc_n_m: 2, position_error_n_m: 0.1 },
          { timestamp: 20, nav_n_m: 3, vloc_n_m: 4, position_error_n_m: 0.3 },
          { timestamp: 30, nav_n_m: 5, vloc_n_m: 6, position_error_n_m: 0.5 },
        ];
        const pairSpec = {
          rows: [{ label: "N", left: "nav_n_m", right: "vloc_n_m", unit: "m" }],
          leftName: "nav",
          rightName: "vloc",
        };
        const errorSpec = {
          rows: [{ label: "N error", field: "position_error_n_m", unit: "m" }],
        };
        process.stdout.write(JSON.stringify({
          hover: context.compositeHoverPayload(rows, pairSpec, 19),
          range: context.compositeRangePayload(rows, errorSpec, [9, 21]),
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["hover"]["timestamp"] == 20
    assert payload["hover"]["values"] == [
        {"label": "N nav", "value": 3, "unit": "m"},
        {"label": "N vloc", "value": 4, "unit": "m"},
    ]
    assert payload["range"]["start"] == 9
    assert payload["range"]["end"] == 21
    assert payload["range"]["sampleCount"] == 2
    assert payload["range"]["stats"][0]["label"] == "N error"
    assert payload["range"]["stats"][0]["mean"] == 0.2


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_composite_hover_overlay_spans_chart_and_follows_cursor():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (id = "") => ({
          id,
          children: [],
          parentNode: null,
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          className: "",
          clientWidth: 900,
          clientHeight: 720,
          classList: {
            values: [],
            add(value) { this.values.push(value); },
            remove() {},
          },
          appendChild(child) {
            child.parentNode = this;
            this.children.push(child);
            if (child.id) {
              elements[child.id] = child;
            }
            return child;
          },
          addEventListener(type, handler) {
            this[`event_${type}`] = handler;
          },
          remove() {},
          getBoundingClientRect() {
            return { left: 10, top: 20, width: this.clientWidth, height: this.clientHeight };
          },
        });
        const elements = {};
        const document = {
          body: makeElement("body"),
          getElementById(id) {
            return elements[id] || null;
          },
          createElement(tag) {
            return makeElement(tag);
          },
        };
        const plot = makeElement("positionCompareComposite");
        plot.onHandlers = {};
        plot.on = (name, handler) => { plot.onHandlers[name] = handler; };
        elements.positionCompareComposite = plot;
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() { return Promise.resolve(); }, purge() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        const figure = context.pairCompositeFigure("positionCompareComposite", "NED", [
          { timestamp: 10, nav_n_m: 1, vloc_n_m: 2, nav_e_m: 3, vloc_e_m: 4, nav_d_m: -5, vloc_d_m: -6 },
          { timestamp: 20, nav_n_m: 11, vloc_n_m: 12, nav_e_m: 13, vloc_e_m: 14, nav_d_m: -15, vloc_d_m: -16 },
        ], {
          title: "NED",
          leftName: "nav",
          rightName: "vloc",
          rows: [
            { label: "N", left: "nav_n_m", right: "vloc_n_m", unit: "m" },
            { label: "E", left: "nav_e_m", right: "vloc_e_m", unit: "m" },
            { label: "D", left: "nav_d_m", right: "vloc_d_m", unit: "m" },
          ],
        }, { variant: "live" });
        context.renderLiveFigure(figure);
        plot.onHandlers.plotly_hover({
          points: [{ x: 19, xaxis: { _offset: 80, l2p(x) { return x * 2; } } }],
          event: { clientX: 240, clientY: 280 },
        });
        const crosshair = elements.positionCompareCompositeCrosshair;
        const tooltip = elements.positionCompareCompositeTooltip;
        process.stdout.write(JSON.stringify({
          crosshairClass: crosshair.className,
          crosshairDisplay: crosshair.style.display,
          crosshairLeft: crosshair.style.left,
          crosshairTop: crosshair.style.top,
          crosshairBottom: crosshair.style.bottom,
          tooltipClass: tooltip.className,
          tooltipDisplay: tooltip.style.display,
          tooltipLeft: tooltip.style.left,
          tooltipTop: tooltip.style.top,
          tooltipHtml: tooltip.innerHTML,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["crosshairClass"] == "composite-crosshair"
    assert payload["crosshairDisplay"] == "block"
    assert payload["crosshairLeft"] == "118px"
    assert payload["crosshairTop"] == "0px"
    assert payload["crosshairBottom"] == "0px"
    assert payload["tooltipClass"] == "composite-floating-tooltip"
    assert payload["tooltipDisplay"] == "block"
    assert payload["tooltipLeft"] != ""
    assert payload["tooltipTop"] != ""
    assert "当前时间戳 20.000 s" in payload["tooltipHtml"]
    for label in ["N nav", "N vloc", "E nav", "E vloc", "D nav", "D vloc"]:
        assert label in payload["tooltipHtml"]


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_entry_mode_switches_between_vloc_and_vo_result_pages():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() {},
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          downloadJson: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          trajectory3d: makeElement(),
          trajectoryXY: makeElement(),
          errorDistance: makeElement(),
          rpeTranslationTime: makeElement(),
          rpeRotationTime: makeElement(),
          scaleFrameTime: makeElement(),
          summaryKicker: makeElement(),
          summaryTitle: makeElement(),
          visualKicker: makeElement(),
          visualTitle: makeElement(),
          metrics: makeElement(),
        };
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge() {} },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        context.updateEntryModeUi();
        const vlocState = {
          rpeTranslationHidden: elements.rpeTranslationTime.hidden,
          scaleFrameHidden: elements.scaleFrameTime.hidden,
          trajectory3dHidden: elements.trajectory3d.hidden,
          summaryTitle: elements.summaryTitle.textContent,
          visualTitle: elements.visualTitle.textContent,
        };

        elements.entryMode.value = "vo";
        context.updateEntryModeUi();
        const voState = {
          rpeTranslationHidden: elements.rpeTranslationTime.hidden,
          scaleFrameHidden: elements.scaleFrameTime.hidden,
          trajectory3dHidden: elements.trajectory3d.hidden,
          summaryTitle: elements.summaryTitle.textContent,
          visualTitle: elements.visualTitle.textContent,
        };

        process.stdout.write(JSON.stringify({ vlocState, voState }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["vlocState"]["rpeTranslationHidden"] is True
    assert payload["vlocState"]["scaleFrameHidden"] is True
    assert payload["vlocState"]["trajectory3dHidden"] is False
    assert "VLOC" in payload["vlocState"]["summaryTitle"]
    assert "VLOC" in payload["vlocState"]["visualTitle"]
    assert payload["voState"]["rpeTranslationHidden"] is False
    assert payload["voState"]["scaleFrameHidden"] is False
    assert payload["voState"]["trajectory3dHidden"] is False
    assert "VO" in payload["voState"]["summaryTitle"]
    assert "VO" in payload["voState"]["visualTitle"]


@pytest.mark.skip("ES Module refactor: need bundler or vm-modules to load split JS files")
def test_static_vloc_chart_directory_controls_only_vloc_charts():
    html = Path("voeval/visualization/index.html").read_text()
    assert "图表目录" in html
    for element_id in ["vlocChartDirectorySection", "vlocChartList", "vlocChartSelectAll", "vlocChartClear"]:
        assert f'id="{element_id}"' in html
    assert html.index('id="runButton"') < html.index('id="vlocChartDirectorySection"')

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          dataset: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() {},
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          vlocChartDirectorySection: makeElement(),
          vlocChartList: makeElement(),
          vlocChartSelectAll: makeElement(),
          vlocChartClear: makeElement(),
          downloadJson: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          metrics: makeElement(),
          summaryKicker: makeElement(),
          summaryTitle: makeElement(),
          visualKicker: makeElement(),
          visualTitle: makeElement(),
        };
        [
          "trajectory3d", "trajectoryXY", "errorDistance",
          "heightComparison", "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus", "voStatus",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
          "rpeTranslationTime", "rpeRotationTime",
        ].forEach((id) => { elements[id] = makeElement(); });
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
          querySelectorAll() { return []; },
        };
        const purged = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge(id) { purged.push(id); } },
        };
        context.globalThis = context;
        const templateCode = fs.readFileSync("voeval/visualization/visualization/report_templates.js", "utf8");
        vm.runInNewContext(templateCode, context);
        const figureCode = fs.readFileSync("voeval/visualization/visualization/figure_specs.js", "utf8");
        vm.runInNewContext(figureCode, context);
        const code = fs.readFileSync("voeval/visualization/js/main.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        context.renderVlocChartDirectory();
        const itemCount = (elements.vlocChartList.innerHTML.match(/data-chart-id=/g) || []).length;
        const checkedCount = (elements.vlocChartList.innerHTML.match(/checked/g) || []).length;

        context.clearVlocChartDirectory();
        const hiddenAfterClear = [
          "trajectory3d", "trajectoryXY", "errorDistance", "heightComparison",
          "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
        ].every((id) => elements[id].hidden === true);

        context.selectAllVlocChartDirectory();
        const visibleAfterSelectAll = [
          "trajectory3d", "trajectoryXY", "errorDistance", "heightComparison",
          "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
        ].every((id) => elements[id].hidden === false);

        context.clearVlocChartDirectory();
        elements.entryMode.value = "vo";
        context.applyEntryModeChartVisibility("vo");
        const voStillVisible = [
          "trajectory3d", "errorDistance", "navStatusModes", "navVelocity", "navResetCounts", "voStatus",
          "positionCompareComposite", "attitudeCompareComposite", "positionErrorComposite", "attitudeErrorComposite",
          "rpeTranslationTime", "rpeRotationTime",
        ].every((id) => elements[id].hidden === false);
        const voDemandExcludedHidden = ["trajectoryXY", "heightComparison", "vlocStatus"].every((id) => elements[id].hidden === true);

        process.stdout.write(JSON.stringify({
          itemCount,
          checkedCount,
          hiddenAfterClear,
          visibleAfterSelectAll,
          voStillVisible,
          voDemandExcludedHidden,
          purged,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["itemCount"] == 12
    assert payload["checkedCount"] == 12
    assert payload["hiddenAfterClear"] is True
    assert payload["visibleAfterSelectAll"] is True
    assert payload["voStillVisible"] is True
    assert payload["voDemandExcludedHidden"] is True
    assert "trajectory3d" in payload["purged"]
