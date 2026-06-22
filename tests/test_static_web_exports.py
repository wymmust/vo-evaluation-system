import base64
import io
import importlib.util
import json
import subprocess
import textwrap
from pathlib import Path

from openpyxl import load_workbook


def test_static_browser_evaluator_exports_new_vloc_summary_metrics():
    source = Path("static_web/py/evaluator.py").read_text()
    assert "mean_error_pos_xy" in source
    assert "mean_error_pos_z" in source
    assert "mean_error_euler" in source
    assert "max_error_pos_xy" in source
    assert "max_error_pos_z" in source
    assert "max_error_euler" in source


def test_static_directory_entry_ui_uses_vloc_vo_modes_instead_of_legacy_file_formats():
    html = Path("static_web/index.html").read_text()
    assert 'id="entryMode"' in html
    assert 'id="dataDirFiles"' in html
    assert 'id="logDirFiles"' in html
    assert 'id="dataDirButton"' in html
    assert 'id="logDirButton"' in html
    assert 'id="dataDirStatus"' in html
    assert 'id="logDirStatus"' in html
    assert "webkitdirectory" in html
    assert 'id="gtFile"' not in html
    assert 'id="estFile"' not in html
    assert 'id="gtFormat"' not in html
    assert 'id="estFormat"' not in html
    assert 'id="modeAndAlignmentSection"' not in html
    assert 'id="alignment"' not in html
    assert 'id="orientationCorrection"' not in html
    assert 'id="associationMode"' not in html
    assert 'id="positionCompareComposite"' in html
    assert 'id="attitudeCompareComposite"' in html
    assert 'id="positionErrorComposite"' in html
    assert 'id="attitudeErrorComposite"' in html
    assert 'id="voChartDirectorySection" data-entry-show="vo"' in html
    assert 'id="voChartList"' in html
    assert 'id="voChartSelectAll"' in html
    assert 'id="voChartClear"' in html
    assert 'id="evaluationParametersSection" data-entry-hide="vloc"' in html
    assert 'id="rpeDeltaValue"' in html
    assert 'id="rpeDeltaUnit"' in html
    assert 'id="scaleDeltaValue"' in html
    assert 'id="scaleDeltaUnit"' in html
    for removed_id in [
        "maxTimeDiff",
        "interpolationPreset",
        "maxInterpolationGap",
        "allowExtrapolation",
        "interpolateRotation",
        "timeOffset",
        "segmentLengths",
        "maxSegments",
        "segmentStep",
        "lengthTolerance",
        "segmentPolicy",
        "discontinuityStep",
        "discontinuityGap",
        "divergenceAbs",
        "divergenceRel",
    ]:
        assert f'id="{removed_id}"' not in html
    assert 'id="seriesXTime"' not in html

    source = Path("static_web/app.js").read_text()
    assert "VO_CHART_OPTIONS" in source
    assert "state.voSelectedChartIds" in source
    assert "renderVoChartDirectory" in source
    assert "selectedVoChartIds" in source
    for chart_id in [
        "trajectoryXY",
        "segmentError",
        "speedError",
        "sim3ScaleTime",
    ]:
        assert f'id: "{chart_id}"' not in source.split("const VO_CHART_OPTIONS = [", 1)[1].split("];", 1)[0]
    for chart_id in [
        "trajectory3d",
        "errorDistance",
        "navStatusModes",
        "navVelocity",
        "navResetCounts",
        "voStatus",
        "positionCompareComposite",
        "attitudeCompareComposite",
        "positionErrorComposite",
        "attitudeErrorComposite",
        "rpeTranslationTime",
        "rpeRotationTime",
        "scaleFrameTime",
    ]:
        assert f'id: "{chart_id}"' in source.split("const VO_CHART_OPTIONS = [", 1)[1].split("];", 1)[0]

    css = Path("static_web/style.css").read_text()
    assert "[hidden]" in css
    assert "display: none !important" in css
    assert "grid-template-columns: repeat(3, minmax(0, 1fr));" in css
    assert "font-size: 13px;" in css


def test_static_dead_report_and_time_series_helpers_are_removed():
    source = Path("static_web/app.js").read_text()
    removed_names = [
        "associationLabel",
        "renderGtVoTimeChart",
        "renderErrorTimeChart",
        "reportSubtitle",
        "reportOverallStatus",
        "buildReportMetricCards",
        "reportMetricCardHtml",
        "buildReportFindings",
        "reportFindingHtml",
        "buildSegmentSummaryRows",
    ]
    for name in removed_names:
        assert f"function {name}" not in source


def test_streamlit_frontend_defaults_align_with_static_web_directory_flow():
    source = Path("app.py").read_text()
    assert 'st.radio("评估入口", list(EVALUATION_ENTRY_OPTIONS), index=0)' in source
    assert 'if entry_mode == "vloc":' in source
    assert 'alignment="none"' in source
    assert 'orientation_correction="none"' in source
    assert 'association_mode="interpolate_gt"' in source
    assert 'max_interpolation_gap_s=1.0' in source
    assert "st.number_input(\"RPE 统计间隔\"" in source
    assert "st.number_input(\"尺度图间隔\"" in source
    for removed_widget_call in [
        'st.text_input("长航程子轨迹长度 m"',
        'st.number_input("每个长度最多抽样段数"',
        'st.number_input("子轨迹起点步长',
        'st.number_input("子轨迹长度容差比例"',
        'st.number_input("断点步长阈值 m"',
        'st.number_input("断点时间间隔阈值 s"',
        'st.number_input("发散绝对阈值 m"',
        'st.number_input("发散相对阈值 % 路程"',
    ]:
        assert removed_widget_call not in source
    assert 'if entry_mode == "vo":' in source
    assert 'report = evaluator.evaluate_vo_bundle(bundle, cfg)' in source
    assert 'segment_policy_label = "按VO连续段逐段评估"' in source
    assert 'entry_mode == "vo" and st.selectbox("轨迹对齐"' not in source
    assert "VO_CHART_OPTIONS" in source
    assert "show_chart_directory(\"vo\", VO_CHART_OPTIONS)" in source
    assert "selected_vo_chart_ids" in source
    assert "show_visuals(report, entry_mode, selected_vloc_chart_ids, selected_vo_chart_ids)" in source
    assert 'segment_policy_label = "按VO时间戳统一评估（推荐）"' in source
    assert "视觉布局与 static_web 保持同一套信息组织" not in source


def test_static_directory_picker_shows_selected_directory_name_in_custom_status():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() { this.clicked = true; },
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
        };
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
        };
        const plotlyCallStats = { addTraceCalls: 0, deleteTraceCalls: 0 };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        elements.dataDirFiles.files = [
          { name: "imu.txt", webkitRelativePath: "data_dir/imu.txt" },
          { name: "extra.txt", webkitRelativePath: "data_dir/extra.txt" },
        ];
        elements.logDirFiles.files = [
          { name: "vloc.txt", webkitRelativePath: "log_dir/vloc.txt" },
        ];

        context.updateDirectoryStatus("data");
        context.updateDirectoryStatus("log");

        process.stdout.write(JSON.stringify({
          dataStatus: elements.dataDirStatus.textContent,
          logStatus: elements.logDirStatus.textContent,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert "data_dir" in payload["dataStatus"]
    assert "2" in payload["dataStatus"]
    assert "log_dir" in payload["logStatus"]
    assert "1" in payload["logStatus"]


def test_static_vloc_mode_hides_transform_controls_and_uses_fixed_sync_defaults():
    html = Path("static_web/index.html").read_text()
    assert 'id="modeAndAlignmentSection"' not in html
    assert 'data-entry-hide="vloc"' in html

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() {},
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          rpeDeltaValue: makeElement("1"),
          rpeDeltaUnit: makeElement("frames"),
          scaleDeltaValue: makeElement("100"),
          scaleDeltaUnit: makeElement("meters"),
        };
        const hiddenNodes = [
          { dataset: { entryHide: "vloc" }, hidden: false },
          { dataset: { entryHide: "vo" }, hidden: false },
        ];
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
          querySelectorAll(selector) {
            if (selector === "[data-entry-hide]") return hiddenNodes;
            return [];
          },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        context.updateEntryModeUi();
        process.stdout.write(JSON.stringify({
          config: context.buildConfig(),
          hiddenStates: hiddenNodes.map((node) => node.hidden),
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["config"]["alignment"] == "none"
    assert payload["config"]["orientation_correction"] == "none"
    assert payload["config"]["association_mode"] == "interpolate_gt"
    assert payload["config"]["max_interpolation_gap_s"] == 1.0
    assert payload["config"]["allow_extrapolation"] is False
    assert payload["config"]["time_offset_s"] == 0.0
    assert payload["config"]["continuous_segment_policy"] == "vo_timestamps"
    assert payload["hiddenStates"] == [True, False]


def test_static_vo_mode_hides_fixed_workflow_controls_and_uses_fixed_defaults():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() {},
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vo"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          rpeDeltaValue: makeElement("100"),
          rpeDeltaUnit: makeElement("meters"),
          scaleDeltaValue: makeElement("50"),
          scaleDeltaUnit: makeElement("meters"),
        };
        const hiddenNodes = [
          { dataset: { entryHide: "vloc,vo" }, hidden: false },
          { dataset: { entryHide: "vloc" }, hidden: false },
        ];
        const shownNodes = [
          { dataset: { entryShow: "vo" }, hidden: true },
          { dataset: { entryShow: "vloc,vo" }, hidden: true },
        ];
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
          querySelectorAll(selector) {
            if (selector === "[data-entry-hide]") return hiddenNodes;
            if (selector === "[data-entry-show]") return shownNodes;
            return [];
          },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        context.updateEntryModeUi();
        process.stdout.write(JSON.stringify({
          config: context.buildConfig(),
          hiddenStates: hiddenNodes.map((node) => node.hidden),
          shownStates: shownNodes.map((node) => node.hidden),
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["config"]["alignment"] == "sim3"
    assert payload["config"]["orientation_correction"] == "none"
    assert payload["config"]["association_mode"] == "interpolate_gt"
    assert payload["config"]["max_time_diff_s"] is None
    assert payload["config"]["max_interpolation_gap_s"] == 1.0
    assert payload["config"]["allow_extrapolation"] is False
    assert payload["config"]["interpolate_rotation"] is True
    assert payload["config"]["time_offset_s"] == 0.0
    assert payload["config"]["continuous_segment_policy"] == "segments"
    assert payload["config"]["rpe_delta_value"] == 100
    assert payload["config"]["rpe_delta_unit"] == "meters"
    assert payload["config"]["scale_delta_value"] == 50
    assert payload["config"]["scale_delta_unit"] == "meters"
    assert payload["config"]["segment_lengths_m"] == [50, 100, 200, 500, 1000, 2000, 5000]
    assert payload["config"]["max_segments_per_length"] == 10000
    assert payload["config"]["segment_step_frames"] == 10
    assert payload["config"]["max_segment_length_diff_ratio"] == 0.05
    assert payload["config"]["discontinuity_step_m"] == 100
    assert payload["config"]["discontinuity_time_gap_s"] == 5
    assert payload["config"]["divergence_abs_m"] == 30
    assert payload["config"]["divergence_rel_percent"] == 3
    assert payload["hiddenStates"] == [True, False]
    assert payload["shownStates"] == [False, False]


def test_static_browser_runner_uses_fixed_bundle_parsers_instead_of_legacy_single_file_loader(tmp_path):
    runner_path = Path("static_web/py/browser_runner.py")
    spec = importlib.util.spec_from_file_location("browser_runner_test", runner_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)

    imu_rows = [
        "ts ts_fcc status flight_mode x y z yaw pitch roll vx vy vz position_reset_count altitude_reset_count heading_reset_count latitude longitude altitude altitude_msl height"
    ]
    for i in range(220):
        t = 10.0 + i * 0.1
        imu_rows.append(
            f"{t:.1f} {100.0 + i * 0.1:.1f} 4194305 3 {t:.6f} 0 0 0 0 0 1 0 0 0 1 2 31.1 121.2 50 51 5"
        )
    imu_text = "\n".join(imu_rows)
    vloc_text = """ts status num_inliers reset_count x y z yaw pitch roll latitude longitude height
10.0 2 42 0 1.1 2.1 3.1 90 2 -1 31.1 121.2 5
10.1 3 43 1 2.1 3.1 4.1 91 3 -2 31.2 121.3 6
"""
    vo_rows = ["ts num_inliers x y z yaw pitch roll is_keyframe time_cost reset_count depth_mean depth_min depth_max"]
    for i in range(201):
        t = 10.0 + i * 0.1
        vo_rows.append(f"{t:.1f} 50 {t:.6f} 0 0 0 0 0 1 12.5 0 4.1 0.2 8.9")
    vo_text = "\n".join(vo_rows)
    home_text = "121.2 31.1 51.0\n"
    calib_text = """%YAML:1.0
---
T_imu_body: [ 1, 0, 0, 0.1, 0, 1, 0, 0.2, 0, 0, 1, 0.3, 0, 0, 0, 1 ]
cam0:
  T_cam_imu: [ 0, -1, 0, 1, 1, 0, 0, 2, 0, 0, 1, 3, 0, 0, 0, 1 ]
"""
    config_json = json.dumps({"segment_lengths_m": [50, 100], "max_interpolation_gap_s": 0.3})

    vloc_report = json.loads(
        module.evaluate_vloc_bundle_json(
            imu_text,
            vloc_text,
            home_text,
            calib_text,
            config_json,
            "imu.txt",
            "vloc.txt",
            "home_point.txt",
            "calib_raw.yaml",
        )
    )
    assert vloc_report["summary"]["matched_poses"] == 2

    vo_report = json.loads(
        module.evaluate_vo_bundle_json(
            imu_text,
            vo_text,
            home_text,
            calib_text,
            config_json,
            "imu.txt",
            "vo.txt",
            "home_point.txt",
            "calib_raw.yaml",
        )
    )
    assert vo_report["summary"]["matched_poses"] == 201
    assert vo_report["association"]["valid_est_after_segment_filter"] == 201

    vloc_light_report = json.loads(
        module.evaluate_vloc_bundle_json_light(
            imu_text,
            vloc_text,
            home_text,
            calib_text,
            config_json,
            "imu.txt",
            "vloc.txt",
            "home_point.txt",
            "calib_raw.yaml",
        )
    )
    assert "scale_per_frame" not in vloc_light_report.get("trajectory_exports", {})

    vo_light_report = json.loads(
        module.evaluate_vo_bundle_json_light(
            imu_text,
            vo_text,
            home_text,
            calib_text,
            config_json,
            "imu.txt",
            "vo.txt",
            "home_point.txt",
            "calib_raw.yaml",
        )
    )
    assert "per_pose" not in vo_light_report
    assert "segment_records" not in vo_light_report
    assert "scale_per_frame" in vo_light_report["trajectory_exports"]
    assert vo_light_report["trajectory_exports"]["scale_per_frame"][0]["scale_available"] is True


def test_static_python_sources_are_fetched_without_browser_cache():
    worker_js = Path("static_web/worker.js").read_text()

    assert 'cache: "no-store"' in worker_js
    assert "cacheBust" in worker_js


def test_static_runtime_uses_worker_and_layered_report_slices():
    app_js = Path("static_web/app.js").read_text()
    worker_js = Path("static_web/worker.js").read_text()
    runner = Path("static_web/py/browser_runner.py").read_text()
    html = Path("static_web/index.html").read_text()

    assert 'new Worker("./worker.js")' in app_js
    assert 'workerRequest("evaluate"' in app_js
    assert 'workerRequest("slice"' in app_js
    assert "loadPyodide" not in app_js
    assert "pyodide.js" not in html
    assert "evaluate_vloc_bundle_json_light" in worker_js
    assert "evaluate_vo_bundle_json_light" in worker_js
    assert "get_report_slice_json" in worker_js
    assert "def _light_report" in runner
    assert '"per_pose", "segment_records", "trajectory_exports"' in runner
    assert "download_slices_available" in runner


def test_static_scale_interval_controls_are_wired_into_config():
    html = Path("static_web/index.html").read_text()
    assert 'id="rpeDeltaValue"' in html
    assert 'id="rpeDeltaUnit"' in html
    assert 'id="scaleDeltaValue"' in html
    assert 'id="scaleDeltaUnit"' in html
    for removed_id in [
        "maxTimeDiff",
        "maxInterpolationGap",
        "allowExtrapolation",
        "interpolateRotation",
        "timeOffset",
        "segmentLengths",
        "maxSegments",
        "segmentStep",
        "lengthTolerance",
        "discontinuityStep",
        "discontinuityGap",
        "divergenceAbs",
        "divergenceRel",
    ]:
        assert f'id="{removed_id}"' not in html

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const elements = {
          entryMode: { value: "vo" },
          rpeDeltaValue: { value: "1" },
          rpeDeltaUnit: { value: "frames" },
          scaleDeltaValue: { value: "100" },
          scaleDeltaUnit: { value: "meters" },
        };
        const element = { addEventListener() {}, classList: { add() {}, remove() {} }, style: {}, files: [], value: "" };
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        process.stdout.write(JSON.stringify(context.buildConfig()));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    config = json.loads(result.stdout)
    assert config["scale_delta_value"] == 100
    assert config["scale_delta_unit"] == "meters"
    assert config["scale_distance_tolerance_ratio"] == 0.05


def test_static_html_export_keeps_light_chart_exports_and_xlsx_has_sheets():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const report = {
          inputs: { entry_mode: "vo" },
          summary: { matched_poses: 2 },
          per_pose: [{ timestamp: 1, error_m: 0.1 }],
            trajectory_exports: {
              input_gt_tum: [{ timestamp: 1, tx: 0, ty: 0, tz: 0 }],
              input_vo_tum: [{ timestamp: 1, tx: 1, ty: 0, tz: 0 }],
              filtered_vo_tum: [{ timestamp: 1, tx: 1, ty: 0, tz: 0 }],
              interpolated_gt_tum: [{ timestamp: 1, tx: 0, ty: 0, tz: 0 }],
              sim3_gt_tum: [{ timestamp: 1, tx: 0, ty: 0, tz: 0 }],
              sim3_vo_tum: [{ timestamp: 1, tx: 0, ty: 0, tz: 0 }],
              ate_per_frame: [{ timestamp: 1, ate_position_m: 0.1 }],
              rpe_per_frame: [{ timestamp: 1, rpe_translation_m: 0.2, rpe_available: true }],
              scale_per_frame: [{ timestamp: 1, local_sim3_scale: 2, scale_available: true }],
            },
          };
            const sanitized = context.reportForHtmlExport(report);
        const vlocSanitized = context.reportForHtmlExport({
          inputs: { entry_mode: "vloc" },
          trajectory_exports: {
            rpe_per_frame: [{ timestamp: 1, rpe_translation_m: 0.2, rpe_available: true }],
            scale_per_frame: [{ timestamp: 1, local_sim3_scale: 2, scale_available: true }],
          },
        });
            const workbook = context.buildTrajectoryWorkbook(report.trajectory_exports);
            process.stdout.write(JSON.stringify({
              sanitized,
          vlocSanitized,
              workbookBase64: Buffer.from(workbook).toString("base64"),
            }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)

    assert payload["sanitized"]["trajectory_exports"] == {
        "rpe_per_frame": [{"timestamp": 1, "rpe_translation_m": 0.2, "rpe_available": True}],
        "scale_per_frame": [{"timestamp": 1, "local_sim3_scale": 2, "scale_available": True}],
    }
    assert "scale_per_frame" not in payload["vlocSanitized"].get("trajectory_exports", {})
    assert "input_gt_tum" not in payload["sanitized"]["trajectory_exports"]
    assert "sim3_vo_tum" not in payload["sanitized"]["trajectory_exports"]
    assert "ate_per_frame" not in payload["sanitized"]["trajectory_exports"]
    assert "per_pose" not in payload["sanitized"]
    assert "segment_records" not in payload["sanitized"]
    assert payload["sanitized"]["summary"]["matched_poses"] == 2

    workbook_bytes = base64.b64decode(payload["workbookBase64"])
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
    assert workbook.sheetnames == [
        "input_gt_tum",
        "input_vo_tum",
        "filtered_vo_tum",
        "interpolated_gt_tum",
        "sim3_gt_tum",
        "sim3_vo_tum",
        "ate_per_frame",
        "rpe_per_frame",
        "scale_per_frame",
    ]
    assert workbook["input_gt_tum"]["A1"].value == "timestamp"
    assert workbook["input_gt_tum"]["A2"].value == 1
    assert workbook["ate_per_frame"]["B1"].value == "ate_position_m"
    assert workbook["rpe_per_frame"]["B1"].value == "rpe_translation_m"
    assert workbook["scale_per_frame"]["B1"].value == "local_sim3_scale"


def test_static_trajectory_workbook_omits_missing_vloc_sim3_sheets():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const workbook = context.buildTrajectoryWorkbook({
          input_gt_tum: [{ timestamp: 1, tx: 0 }],
          input_vo_tum: [{ timestamp: 1, tx: 1 }],
          filtered_vo_tum: [{ timestamp: 1, tx: 1 }],
          interpolated_gt_tum: [{ timestamp: 1, tx: 0 }],
          ate_per_frame: [{ timestamp: 1, ate_position_m: 0.1 }],
          rpe_per_frame: [{ timestamp: 1, rpe_translation_m: 0.2 }],
        });
        process.stdout.write(Buffer.from(workbook).toString("base64"));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    workbook = load_workbook(io.BytesIO(base64.b64decode(result.stdout)), read_only=True)
    assert workbook.sheetnames == [
        "input_gt_tum",
        "input_vo_tum",
        "filtered_vo_tum",
        "interpolated_gt_tum",
        "ate_per_frame",
        "rpe_per_frame",
    ]


def test_static_export_filenames_include_directory_name_and_entry_mode():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "vloc",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const vlocReport = {
          inputs: {
            entry_mode: "vloc",
            data_dir_name: "2839_traj",
            log_dir_name: "2839_traj",
          },
        };
        const vlocHtml = context.evaluationExportFilename("evaluation_report", "html", vlocReport);
        const vlocXlsx = context.evaluationExportFilename("trajectory_exports", "xlsx", vlocReport);
        const voReport = {
          inputs: {
            entry_mode: "vo",
            data_dir_name: "flight:data",
            log_dir_name: "log/vo*run",
          },
        };
        const voHtml = context.evaluationExportFilename("evaluation_report", "html", voReport);
        const defaultDirReport = {
          inputs: {
            entry_mode: "vloc",
            data_dir_name: "data_dir",
            log_dir_name: "log_dir",
          },
        };
        const defaultDirHtml = context.evaluationExportFilename("evaluation_report", "html", defaultDirReport);
        process.stdout.write(JSON.stringify({ vlocHtml, vlocXlsx, voHtml, defaultDirHtml }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["vlocHtml"] == "2839_traj_vloc_evaluation_report.html"
    assert payload["vlocXlsx"] == "2839_traj_vloc_trajectory_exports.xlsx"
    assert payload["voHtml"] == "flight_data__log_vo_run_vo_evaluation_report.html"
    assert payload["defaultDirHtml"] == "vloc_evaluation_report.html"


def test_static_html_export_is_visualization_snapshot_with_point_selection():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "vloc",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const report = {
          inputs: { entry_mode: "vloc" },
          summary: { matched_poses: 2, gt_path_length_m: 10, duration_s: 1, est_pose_coverage_ratio: 1, gt_pose_coverage_ratio: 1 },
          ate_position_m: { rmse: 0.2 },
          ate_vertical_m: { rmse: 0.1 },
          discontinuities: { all_matches: { break_count: 0 }, selected_segment: { policy: "vo_timestamps" } },
          vloc_details: {
            summary: { mean_error_pos_xy: 0.1, mean_error_pos_z: 0.2, mean_error_euler: 0.3, max_error_pos_xy: 0.4, max_error_pos_z: 0.5, max_error_euler: 0.6 },
            comparison: [
              { timestamp: 1, distance_m: 0, visual_segment_id: 0, nav_n_m: 0, nav_e_m: 0, nav_d_m: 0, vloc_n_m: 0.1, vloc_e_m: 0.1, vloc_d_m: 0.1, nav_height_m: 5, vloc_height_m: 4.9, position_error_3d_m: 0.2, horizontal_position_error_m: 0.1, vertical_position_error_abs_m: 0.1, position_error_n_m: 0.1, position_error_e_m: 0.1, position_error_d_m: 0.1, nav_yaw_deg: 0, vloc_yaw_deg: 0.1, nav_pitch_deg: 0, vloc_pitch_deg: 0.2, nav_roll_deg: 0, vloc_roll_deg: 0.3, attitude_error_yaw_deg: 0.1, attitude_error_pitch_deg: 0.2, attitude_error_roll_deg: 0.3 },
              { timestamp: 2, distance_m: 1, visual_segment_id: 0, nav_n_m: 1, nav_e_m: 1, nav_d_m: 0, vloc_n_m: 1.1, vloc_e_m: 1.1, vloc_d_m: 0.1, nav_height_m: 6, vloc_height_m: 5.9, position_error_3d_m: 0.2, horizontal_position_error_m: 0.1, vertical_position_error_abs_m: 0.1, position_error_n_m: 0.1, position_error_e_m: 0.1, position_error_d_m: 0.1, nav_yaw_deg: 1, vloc_yaw_deg: 1.1, nav_pitch_deg: 1, vloc_pitch_deg: 1.2, nav_roll_deg: 1, vloc_roll_deg: 1.3, attitude_error_yaw_deg: 0.1, attitude_error_pitch_deg: 0.2, attitude_error_roll_deg: 0.3 },
            ],
            nav_status: [{ timestamp: 1, flight_mode: 3, navi_mode: 5, rtk_yaw: 1, rtk_alti: 0, vx: 1, vy: 2, vz: 3, velocity_norm: 3.7, position_reset_count: 0, altitude_reset_count: 0, heading_reset_count: 0 }],
            vloc_status: [{ timestamp: 1, vloc_mode: 2, num_inliers: 40, reset_count: 0 }],
          },
        };
        process.stdout.write(context.buildHtmlReport(report));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    html = result.stdout

    assert "VLOC 评估结果" in html
    assert "离线可视化快照" in html
    assert "图表目录" in html
    assert "pointSelectionOutput" in html
    assert "togglePointMode" in html
    assert "recordPointSelection" in html
    assert "composite-crosshair" in html
    assert "attachExportCompositeOverlay" in html
    assert 'node.on("plotly_hover"' in html
    assert "exportAllHoverChips" in html
    assert "nearestExportTracePoint" in html
    assert "focusExportPointSelectionFromEvent" in html
    assert "deleteFocusedExportPointSelection" in html
    assert "point.pointNumber" in html
    assert "removeExportSelectionMarkerTraces" in html
    assert "refreshExportChartSelectionMarkers" in html
    assert "keydown" in html
    assert "trajectory3d" in html
    assert "positionCompareComposite" in html
    assert "调参结论摘要" not in html
    assert "高级详情 / 完整指标" not in html
    assert "原始 JSON 指标" not in html

    behavior_script = textwrap.dedent(
        f"""
        const vm = require("vm");
        const html = {json.dumps(html)};
        const scriptStart = html.lastIndexOf("<script>");
        const scriptEnd = html.lastIndexOf("</script>");
        const inline = html.slice(scriptStart + "<script>".length, scriptEnd).replace(/\\ninitExportPage\\(\\);\\n/, "\\n");
        const chartId = "positionCompareComposite";
        const chart = {{
          id: chartId,
          data: [{{ type: "scatter", name: "N nav", x: [1], y: [2] }}],
          style: {{}},
          classList: {{ toggle() {{}} }},
          dataset: {{ chartId }},
        }};
        const pointSelectionOutputSection = {{ hidden: true }};
        const pointSelectionOutput = {{ innerHTML: "" }};
        const document = {{
          getElementById(id) {{
            if (id === chartId) return chart;
            if (id === "pointSelectionOutputSection") return pointSelectionOutputSection;
            if (id === "pointSelectionOutput") return pointSelectionOutput;
            return null;
          }},
          querySelector() {{ return null; }},
          querySelectorAll() {{ return []; }},
        }};
        const Plotly = {{
          addTraces(id, traces) {{
            if (id !== chartId) throw new Error("unexpected addTraces chart " + id);
            chart.data.push(...traces);
          }},
          deleteTraces(id, indices) {{
            if (id !== chartId) throw new Error("unexpected deleteTraces chart " + id);
            [...indices].sort((left, right) => right - left).forEach((index) => chart.data.splice(index, 1));
          }},
        }};
        const context = {{ document, window: {{}}, Plotly, console, Number, String, Array, Math, Set, Date }};
        context.globalThis = context;
        vm.runInNewContext(inline, context);
        const selection = {{
          id: "s1",
          order: 1,
          colorSlot: 1,
          chartId,
          chartTitle: "NED 随时间变化",
          traceName: "N nav",
          timestamp: 1,
          value: "2.000",
          color: "#ef4444",
          markerText: "",
          x: 1,
          y: 2,
          xaxis: "x",
          yaxis: "y",
        }};
        context.window.__VO_EXPORT_SELECTIONS__ = [selection];
        context.window.__VO_EXPORT_FOCUSED_SELECTION_ID__ = "s1";
        context.refreshExportChartSelectionMarkers(chartId);
        const afterAdd = chart.data.length;
        context.deleteFocusedExportPointSelection();
        const afterDelete = chart.data.length;
        context.window.__VO_EXPORT_SELECTIONS__ = [selection];
        context.refreshExportChartSelectionMarkers(chartId);
        const afterReadd = chart.data.length;
        context.clearChartSelections(chartId);
        const afterClear = chart.data.length;
        process.stdout.write(JSON.stringify({{
          afterAdd,
          afterDelete,
          afterReadd,
          afterClear,
          selections: context.window.__VO_EXPORT_SELECTIONS__.length,
        }}));
        """
    )
    behavior_result = subprocess.run(["node", "-e", behavior_script], check=True, capture_output=True, text=True)
    behavior = json.loads(behavior_result.stdout)
    assert behavior == {"afterAdd": 3, "afterDelete": 1, "afterReadd": 3, "afterClear": 1, "selections": 0}


def test_static_html_export_uses_vo_chart_set_for_vo_reports():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "vo",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        const report = {
          inputs: { entry_mode: "vo" },
          summary: { matched_poses: 2, gt_path_length_m: 10, duration_s: 1, est_pose_coverage_ratio: 1, gt_pose_coverage_ratio: 1 },
          ate_position_m: { rmse: 0.2 },
          ate_vertical_m: { rmse: 0.1 },
          rpe_frame_delta: { delta_unit: "frames", delta_frames: 1, translation_m: { rmse: 0.3 } },
          discontinuities: { all_matches: { break_count: 0 }, selected_segment: { policy: "vo_timestamps" } },
          alignment: { scale: 1 },
          vo_details: {
            comparison: [
              { timestamp: 1, distance_m: 0, visual_segment_id: 0, nav_x_m: 0, nav_y_m: 0, nav_z_m: 0, vo_x_aligned_m: 0.1, vo_y_aligned_m: 0.1, vo_z_aligned_m: 0.1, position_error_3d_m: 0.2, horizontal_position_error_m: 0.1, position_error_x_m: 0.1, position_error_y_m: 0.1, position_error_z_m: 0.1, nav_yaw_deg: 0, vo_yaw_aligned_deg: 0.1, nav_pitch_deg: 0, vo_pitch_aligned_deg: 0.2, nav_roll_deg: 0, vo_roll_aligned_deg: 0.3, attitude_error_yaw_deg: 0.1, attitude_error_pitch_deg: 0.2, attitude_error_roll_deg: 0.3 },
              { timestamp: 2, distance_m: 1, visual_segment_id: 0, nav_x_m: 1, nav_y_m: 1, nav_z_m: 0, vo_x_aligned_m: 1.1, vo_y_aligned_m: 1.1, vo_z_aligned_m: 0.1, position_error_3d_m: 0.2, horizontal_position_error_m: 0.1, position_error_x_m: 0.1, position_error_y_m: 0.1, position_error_z_m: 0.1, nav_yaw_deg: 1, vo_yaw_aligned_deg: 1.1, nav_pitch_deg: 1, vo_pitch_aligned_deg: 1.2, nav_roll_deg: 1, vo_roll_aligned_deg: 1.3, attitude_error_yaw_deg: 0.1, attitude_error_pitch_deg: 0.2, attitude_error_roll_deg: 0.3 },
            ],
            nav_status: [{ timestamp: 1, flight_mode: 3, navi_mode: 5, rtk_yaw: 1, rtk_alti: 0, vx: 1, vy: 2, vz: 3, velocity_norm: 3.7, position_reset_count: 0, altitude_reset_count: 0, heading_reset_count: 0 }],
            vo_status: [{ timestamp: 1, num_inliers: 40, is_keyframe: 1, time_cost: 3, reset_count: 0 }],
          },
          trajectory_exports: {
            rpe_per_frame: [{ timestamp: 1, rpe_translation_m: 0.3, rpe_rotation_deg: 0.4, rpe_available: true }],
            scale_per_frame: [{ timestamp: 1, local_sim3_scale: 2, scale_available: true }],
          },
        };
        process.stdout.write(context.buildHtmlReport(report));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    html = result.stdout

    assert "VO 评估结果" in html
    assert "voStatus" in html
    assert "rpeTranslationTime" in html
    assert "rpeRotationTime" in html
    assert "scaleFrameTime" in html
    assert "局部 Sim3 尺度随时间变化" in html
    assert "vlocStatus" not in html
    assert "VLOC 评估结果" not in html


def test_static_visualization_renders_time_series_and_rpe_charts():
    html = Path("static_web/index.html").read_text()
    expected_ids = [
        "trajectory3d",
        "errorDistance",
        "navStatusModes",
        "navVelocity",
        "navResetCounts",
        "voStatus",
        "positionCompareComposite",
        "attitudeCompareComposite",
        "positionErrorComposite",
        "attitudeErrorComposite",
        "rpeTranslationTime",
        "rpeRotationTime",
        "scaleFrameTime",
    ]
    for chart_id in expected_ids:
        assert f'id="{chart_id}"' in html

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const plots = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); }, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        const perPose = [0, 1, 2].map((index) => ({
          timestamp: index,
          segment_id: 0,
          visual_segment_id: index < 2 ? 0 : 1,
          gt_x_m: index,
          gt_y_m: index + 1,
          gt_z_m: index + 2,
          est_x_aligned_m: index + 0.1,
          est_y_aligned_m: index + 1.1,
          est_z_aligned_m: index + 2.1,
          x_error_m: 0.1,
          y_error_m: 0.1,
          z_error_m: 0.1,
          gt_yaw_deg: index,
          gt_pitch_deg: index + 2,
          gt_roll_deg: index + 4,
          est_yaw_aligned_deg: index + 0.2,
          est_pitch_aligned_deg: index + 2.2,
          est_roll_aligned_deg: index + 4.2,
          yaw_error_signed_deg: 0.2,
          pitch_error_signed_deg: 0.2,
          roll_error_signed_deg: 0.2,
          distance_m: index,
          error_m: 0.1,
          horizontal_error_m: 0.1,
          vertical_error_m: 0.1,
        }));
        context.renderCharts({
          inputs: { entry_mode: "vo" },
          per_pose: perPose,
          segment_errors: [],
          speed_bins: [],
          vo_details: {
            nav_status: [
              { timestamp: 0, flight_mode: 3, navi_mode: 5, rtk_yaw: 1, rtk_alti: 0, position_reset_count: 0, altitude_reset_count: 1, heading_reset_count: 2, vx: 0.1, vy: 0.2, vz: 0.3, velocity_norm: 0.374 },
              { timestamp: 1, flight_mode: 4, navi_mode: 6, rtk_yaw: 1, rtk_alti: 1, position_reset_count: 0, altitude_reset_count: 1, heading_reset_count: 3, vx: 0.4, vy: 0.5, vz: 0.6, velocity_norm: 0.877 },
            ],
            vo_status: [
              { timestamp: 0, num_inliers: 40, is_keyframe: 1, time_cost: 3.5, reset_count: 0 },
              { timestamp: 1, num_inliers: 41, is_keyframe: 0, time_cost: 3.7, reset_count: 1 },
            ],
          },
          trajectory_exports: {
            sim3_vo_tum: [
              { timestamp: 0, segment_id: 0, sim3_scale: 2 },
              { timestamp: 1, segment_id: 0, sim3_scale: 2 },
              { timestamp: 2, segment_id: 0, sim3_scale: 2 },
            ],
            rpe_per_frame: [
              { timestamp: 0, rpe_translation_m: 0.3, rpe_rotation_deg: 1.1, rpe_available: true },
              { timestamp: 1, rpe_translation_m: 0.4, rpe_rotation_deg: 1.2, rpe_available: true },
            ],
            scale_per_frame: [
              { timestamp: 0, segment_id: 0, local_sim3_scale: 2, scale_available: true },
              { timestamp: 1, segment_id: 0, local_sim3_scale: 2, scale_available: true },
            ],
          },
        });
        const byId = Object.fromEntries(plots.map((plot) => [plot.id, plot]));
        process.stdout.write(JSON.stringify({
          ids: plots.map((plot) => plot.id),
          trajectory3dNames: byId.trajectory3d.data.map((trace) => trace.name),
          voStartText: byId.trajectory3d.data.find((trace) => trace.name === "vo start").text,
          voEndText: byId.trajectory3d.data.find((trace) => trace.name === "vo end").text,
          voStartMarker: byId.trajectory3d.data.find((trace) => trace.name === "vo start").marker,
          voStartTextFont: byId.trajectory3d.data.find((trace) => trace.name === "vo start").textfont,
          navVelocityNames: byId.navVelocity.data.map((trace) => trace.name),
          voStatusNames: byId.voStatus.data.map((trace) => trace.name),
          positionCompareNames: byId.positionCompareComposite.data.map((trace) => trace.name),
          thirdRowColors: [
            byId.positionCompareComposite.data[4].line.color,
            byId.positionCompareComposite.data[5].line.color,
          ],
          scaleNames: byId.scaleFrameTime.data.map((trace) => trace.name),
          scaleValues: byId.scaleFrameTime.data[0].y,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    rendered_ids = set(payload["ids"])
    assert set(expected_ids).issubset(rendered_ids)
    assert {"trajectoryXY", "segmentError", "speedError", "sim3ScaleTime"}.isdisjoint(rendered_ids)
    assert "GT start" not in payload["trajectory3dNames"]
    assert "GT end" not in payload["trajectory3dNames"]
    assert {"vo start", "vo end"}.issubset(set(payload["trajectory3dNames"]))
    assert payload["voStartText"] == ["vo S1", "vo S2"]
    assert payload["voEndText"] == ["vo E1", "vo E2"]
    assert payload["voStartMarker"]["size"] == 5
    assert payload["voStartMarker"]["color"] == "#9333ea"
    assert payload["voStartMarker"]["line"]["width"] == 1
    assert payload["voStartTextFont"]["size"] == 10
    assert "velocity_norm" in payload["navVelocityNames"]
    assert {"num_inliers", "is_keyframe", "time_cost", "reset_count"}.issubset(set(payload["voStatusNames"]))
    assert {"X Ground truth", "X VO aligned", "Y Ground truth", "Y VO aligned", "Z Ground truth", "Z VO aligned"} == set(payload["positionCompareNames"])
    assert payload["thirdRowColors"] == ["#dc2626", "#0891b2"]
    assert payload["scaleNames"] == ["local_sim3_scale"]
    assert payload["scaleValues"] == [2, 2]


def test_static_composite_angle_time_series_unwraps_180_degree_boundary():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const plots = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); }, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        context.renderPairCompositeChart("attitudeCompareComposite", [
          { timestamp: 0, segment_id: 0, gt_roll_deg: 0, est_roll_aligned_deg: 179 },
          { timestamp: 1, segment_id: 0, gt_roll_deg: 1, est_roll_aligned_deg: -179 },
          { timestamp: 2, segment_id: 0, gt_roll_deg: 2, est_roll_aligned_deg: 178 },
        ], {
          title: "Roll",
          leftName: "Ground truth",
          rightName: "VO aligned",
          rows: [{ label: "Roll", left: "gt_roll_deg", right: "est_roll_aligned_deg", unit: "deg", unwrap: true }],
        });
        process.stdout.write(JSON.stringify(plots[0].data[1].y));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    assert json.loads(result.stdout) == [179, 181, 178]


def test_static_composite_angle_error_time_series_unwraps_180_degree_boundary():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {} }; },
        };
        const plots = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); }, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        context.renderSingleCompositeChart("attitudeErrorComposite", [
          { timestamp: 0, segment_id: 0, roll_error_signed_deg: 179 },
          { timestamp: 1, segment_id: 0, roll_error_signed_deg: -179 },
          { timestamp: 2, segment_id: 0, roll_error_signed_deg: 178 },
        ], {
          title: "Roll error",
          rows: [{ label: "Roll error", field: "roll_error_signed_deg", unit: "deg", unwrap: true }],
        });
        process.stdout.write(JSON.stringify(plots[0].data[0].y));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    assert json.loads(result.stdout) == [179, 181, 178]


def test_static_composite_charts_disable_native_spikes_for_custom_overlay():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {}, appendChild() {} }; },
        };
        const plots = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); return Promise.resolve(); }, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        context.renderPairCompositeChart("positionCompareComposite", [
          { timestamp: 1, nav_n_m: 10, vloc_n_m: 9, nav_e_m: 20, vloc_e_m: 19 },
          { timestamp: 2, nav_n_m: 11, vloc_n_m: 10, nav_e_m: 21, vloc_e_m: 20 },
        ], {
          title: "NED",
          leftName: "nav",
          rightName: "vloc",
          rows: [
            { label: "N", left: "nav_n_m", right: "vloc_n_m", unit: "m" },
            { label: "E", left: "nav_e_m", right: "vloc_e_m", unit: "m" },
          ],
        });
        const plot = plots[0];
        process.stdout.write(JSON.stringify({
          hovermode: plot.layout.hovermode,
          hoversubplots: plot.layout.hoversubplots,
          xaxisShowspikes: plot.layout.xaxis.showspikes,
          xaxis2Showspikes: plot.layout.xaxis2.showspikes,
          firstTraceHoverinfo: plot.data[0].hoverinfo,
          secondTraceHoverinfo: plot.data[1].hoverinfo,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload == {
        "hovermode": "x unified",
        "hoversubplots": "axis",
        "xaxisShowspikes": False,
        "xaxis2Showspikes": False,
        "firstTraceHoverinfo": "none",
        "secondTraceHoverinfo": "none",
    }


def test_static_composite_payload_helpers_compute_hover_and_range_payloads():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const element = {
          addEventListener() {},
          classList: { add() {}, remove() {} },
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          parentNode: { insertBefore() {} },
        };
        const document = {
          body: { appendChild() {} },
          getElementById() { return element; },
          createElement() { return { ...element, click() {}, remove() {}, appendChild() {}, parentNode: element.parentNode }; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() { return Promise.resolve(); }, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        const rows = [
          { timestamp: 10, nav_n_m: 1, vloc_n_m: 2, position_error_n_m: 0.1 },
          { timestamp: 20, nav_n_m: 3, vloc_n_m: 4, position_error_n_m: 0.3 },
          { timestamp: 30, nav_n_m: 5, vloc_n_m: 6, position_error_n_m: 0.5 },
        ];
        const pairSpec = {
          rows: [{ label: "N", left: "nav_n_m", right: "vloc_n_m", unit: "m" }],
          leftName: "nav",
          rightName: "vloc",
        };
        const errorSpec = {
          rows: [{ label: "N error", field: "position_error_n_m", unit: "m" }],
        };
        process.stdout.write(JSON.stringify({
          hover: context.compositeHoverPayload(rows, pairSpec, 19),
          range: context.compositeRangePayload(rows, errorSpec, [9, 21]),
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["hover"]["timestamp"] == 20
    assert payload["hover"]["values"] == [
        {"label": "N nav", "value": 3, "unit": "m"},
        {"label": "N vloc", "value": 4, "unit": "m"},
    ]
    assert payload["range"]["start"] == 9
    assert payload["range"]["end"] == 21
    assert payload["range"]["sampleCount"] == 2
    assert payload["range"]["stats"][0]["label"] == "N error"
    assert payload["range"]["stats"][0]["mean"] == 0.2


def test_static_composite_hover_overlay_spans_chart_and_follows_cursor():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (id = "") => ({
          id,
          children: [],
          parentNode: null,
          style: {},
          files: [],
          value: "",
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          className: "",
          clientWidth: 900,
          clientHeight: 720,
          classList: {
            values: [],
            add(value) { this.values.push(value); },
            remove() {},
          },
          appendChild(child) {
            child.parentNode = this;
            this.children.push(child);
            if (child.id) {
              elements[child.id] = child;
            }
            return child;
          },
          addEventListener(type, handler) {
            this[`event_${type}`] = handler;
          },
          remove() {},
          getBoundingClientRect() {
            return { left: 10, top: 20, width: this.clientWidth, height: this.clientHeight };
          },
        });
        const elements = {};
        const document = {
          body: makeElement("body"),
          getElementById(id) {
            return elements[id] || null;
          },
          createElement(tag) {
            return makeElement(tag);
          },
        };
        const plot = makeElement("positionCompareComposite");
        plot.onHandlers = {};
        plot.on = (name, handler) => { plot.onHandlers[name] = handler; };
        elements.positionCompareComposite = plot;
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() { return Promise.resolve(); }, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        context.renderPairCompositeChart("positionCompareComposite", [
          { timestamp: 10, nav_n_m: 1, vloc_n_m: 2, nav_e_m: 3, vloc_e_m: 4, nav_d_m: -5, vloc_d_m: -6 },
          { timestamp: 20, nav_n_m: 11, vloc_n_m: 12, nav_e_m: 13, vloc_e_m: 14, nav_d_m: -15, vloc_d_m: -16 },
        ], {
          title: "NED",
          leftName: "nav",
          rightName: "vloc",
          rows: [
            { label: "N", left: "nav_n_m", right: "vloc_n_m", unit: "m" },
            { label: "E", left: "nav_e_m", right: "vloc_e_m", unit: "m" },
            { label: "D", left: "nav_d_m", right: "vloc_d_m", unit: "m" },
          ],
        });
        plot.onHandlers.plotly_hover({
          points: [{ x: 19, xaxis: { _offset: 80, l2p(x) { return x * 2; } } }],
          event: { clientX: 240, clientY: 280 },
        });
        const crosshair = elements.positionCompareCompositeCrosshair;
        const tooltip = elements.positionCompareCompositeTooltip;
        process.stdout.write(JSON.stringify({
          crosshairClass: crosshair.className,
          crosshairDisplay: crosshair.style.display,
          crosshairLeft: crosshair.style.left,
          crosshairTop: crosshair.style.top,
          crosshairBottom: crosshair.style.bottom,
          tooltipClass: tooltip.className,
          tooltipDisplay: tooltip.style.display,
          tooltipLeft: tooltip.style.left,
          tooltipTop: tooltip.style.top,
          tooltipHtml: tooltip.innerHTML,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["crosshairClass"] == "composite-crosshair"
    assert payload["crosshairDisplay"] == "block"
    assert payload["crosshairLeft"] == "118px"
    assert payload["crosshairTop"] == "0px"
    assert payload["crosshairBottom"] == "0px"
    assert payload["tooltipClass"] == "composite-floating-tooltip"
    assert payload["tooltipDisplay"] == "block"
    assert payload["tooltipLeft"] != ""
    assert payload["tooltipTop"] != ""
    assert "当前时间戳 20.000 s" in payload["tooltipHtml"]
    for label in ["N nav", "N vloc", "E nav", "E vloc", "D nav", "D vloc"]:
        assert label in payload["tooltipHtml"]


def test_static_entry_mode_switches_between_vloc_and_vo_result_pages():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() {},
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          trajectory3d: makeElement(),
          trajectoryXY: makeElement(),
          errorDistance: makeElement(),
          rpeTranslationTime: makeElement(),
          rpeRotationTime: makeElement(),
          scaleFrameTime: makeElement(),
          summaryKicker: makeElement(),
          summaryTitle: makeElement(),
          visualKicker: makeElement(),
          visualTitle: makeElement(),
          metrics: makeElement(),
        };
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        context.updateEntryModeUi();
        const vlocState = {
          rpeTranslationHidden: elements.rpeTranslationTime.hidden,
          scaleFrameHidden: elements.scaleFrameTime.hidden,
          trajectory3dHidden: elements.trajectory3d.hidden,
          summaryTitle: elements.summaryTitle.textContent,
          visualTitle: elements.visualTitle.textContent,
        };

        elements.entryMode.value = "vo";
        context.updateEntryModeUi();
        const voState = {
          rpeTranslationHidden: elements.rpeTranslationTime.hidden,
          scaleFrameHidden: elements.scaleFrameTime.hidden,
          trajectory3dHidden: elements.trajectory3d.hidden,
          summaryTitle: elements.summaryTitle.textContent,
          visualTitle: elements.visualTitle.textContent,
        };

        process.stdout.write(JSON.stringify({ vlocState, voState }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["vlocState"]["rpeTranslationHidden"] is True
    assert payload["vlocState"]["scaleFrameHidden"] is True
    assert payload["vlocState"]["trajectory3dHidden"] is False
    assert "VLOC" in payload["vlocState"]["summaryTitle"]
    assert "VLOC" in payload["vlocState"]["visualTitle"]
    assert payload["voState"]["rpeTranslationHidden"] is False
    assert payload["voState"]["scaleFrameHidden"] is False
    assert payload["voState"]["trajectory3dHidden"] is False
    assert "VO" in payload["voState"]["summaryTitle"]
    assert "VO" in payload["voState"]["visualTitle"]


def test_static_vloc_chart_directory_controls_only_vloc_charts():
    html = Path("static_web/index.html").read_text()
    assert "图表目录" in html
    for element_id in ["vlocChartDirectorySection", "vlocChartList", "vlocChartSelectAll", "vlocChartClear"]:
        assert f'id="{element_id}"' in html
    assert html.index('id="runButton"') < html.index('id="vlocChartDirectorySection"')

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          dataset: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() {},
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          vlocChartDirectorySection: makeElement(),
          vlocChartList: makeElement(),
          vlocChartSelectAll: makeElement(),
          vlocChartClear: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          metrics: makeElement(),
          summaryKicker: makeElement(),
          summaryTitle: makeElement(),
          visualKicker: makeElement(),
          visualTitle: makeElement(),
        };
        [
          "trajectory3d", "trajectoryXY", "errorDistance",
          "heightComparison", "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus", "voStatus",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
          "rpeTranslationTime", "rpeRotationTime",
        ].forEach((id) => { elements[id] = makeElement(); });
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
          querySelectorAll() { return []; },
        };
        const purged = [];
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge(id) { purged.push(id); } },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        context.renderVlocChartDirectory();
        const itemCount = (elements.vlocChartList.innerHTML.match(/data-chart-id=/g) || []).length;
        const checkedCount = (elements.vlocChartList.innerHTML.match(/checked/g) || []).length;

        context.clearVlocChartDirectory();
        const hiddenAfterClear = [
          "trajectory3d", "trajectoryXY", "errorDistance", "heightComparison",
          "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
        ].every((id) => elements[id].hidden === true);

        context.selectAllVlocChartDirectory();
        const visibleAfterSelectAll = [
          "trajectory3d", "trajectoryXY", "errorDistance", "heightComparison",
          "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
        ].every((id) => elements[id].hidden === false);

        context.clearVlocChartDirectory();
        elements.entryMode.value = "vo";
        context.applyEntryModeChartVisibility("vo");
        const voStillVisible = [
          "trajectory3d", "errorDistance", "navStatusModes", "navVelocity", "navResetCounts", "voStatus",
          "positionCompareComposite", "attitudeCompareComposite", "positionErrorComposite", "attitudeErrorComposite",
          "rpeTranslationTime", "rpeRotationTime",
        ].every((id) => elements[id].hidden === false);
        const voDemandExcludedHidden = ["trajectoryXY", "heightComparison", "vlocStatus"].every((id) => elements[id].hidden === true);

        process.stdout.write(JSON.stringify({
          itemCount,
          checkedCount,
          hiddenAfterClear,
          visibleAfterSelectAll,
          voStillVisible,
          voDemandExcludedHidden,
          purged,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["itemCount"] == 12
    assert payload["checkedCount"] == 12
    assert payload["hiddenAfterClear"] is True
    assert payload["visibleAfterSelectAll"] is True
    assert payload["voStillVisible"] is True
    assert payload["voDemandExcludedHidden"] is True
    assert "trajectory3d" in payload["purged"]


def test_static_vloc_point_selection_excludes_3d_and_records_points():
    html = Path("static_web/index.html").read_text()
    assert 'id="pointSelectionOutputSection"' in html
    assert 'id="pointSelectionOutput"' in html
    assert 'id="clearAllPointSelections"' in html

    css = Path("static_web/style.css").read_text()
    assert ".chart-point-tools" in css
    assert ".point-selection-card" in css
    assert ".selection-point-token" in css

    source = Path("static_web/app.js").read_text()
    assert 'PICKABLE_VLOC_CHART_IDS = VLOC_VISIBLE_CHART_IDS.filter((id) => id !== "trajectory3d")' in source
    assert 'chart.on("plotly_click"' in source
    assert 'chart.on("plotly_hover"' in source
    assert 'event.key !== "Delete"' in source

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          dataset: {},
          data: [],
          classList: { add() {}, remove() {}, toggle() {} },
          addEventListener() {},
          appendChild() {},
          querySelector() { return null; },
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          vlocChartDirectorySection: makeElement(),
          vlocChartList: makeElement(),
          vlocChartSelectAll: makeElement(),
          vlocChartClear: makeElement(),
          pointSelectionOutputSection: makeElement(),
          pointSelectionOutput: makeElement(),
          clearAllPointSelections: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          metrics: makeElement(),
          summaryKicker: makeElement(),
          summaryTitle: makeElement(),
          visualKicker: makeElement(),
          visualTitle: makeElement(),
        };
        [
          "trajectory3d", "trajectoryXY", "errorDistance",
          "heightComparison", "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
          "rpeTranslationTime", "rpeRotationTime",
        ].forEach((id) => { elements[id] = makeElement(); });
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return makeElement(); },
          addEventListener() {},
          querySelectorAll() { return []; },
        };
        const plotlyCallStats = { addTraceCalls: 0, deleteTraceCalls: 0 };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: {
            addTraces(id, traces) { plotlyCallStats.addTraceCalls += 1; elements[id].data.push(...traces); },
            deleteTraces(id, indices) { plotlyCallStats.deleteTraceCalls += 1; elements[id].data = elements[id].data.filter((_trace, index) => !indices.includes(index)); },
            purge(id) { elements[id].data = []; },
          },
          elements,
          plotlyCallStats,
          process,
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(`${code}
          state.report = { inputs: { entry_mode: "vloc" } };
          state.activePointSelectionChartId = "trajectory3d";
          handlePlotPointClick("trajectory3d", { points: [{ x: 1, y: 2, curveNumber: 0, pointNumber: 0, data: { name: "vloc", customdata: [188.5] } }] });
          const after3d = state.pointSelections.length;
          state.activePointSelectionChartId = "trajectoryXY";
          handlePlotPointClick("trajectoryXY", { points: [{ x: 10, y: 20, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [188.5], xaxis: "x", yaxis: "y" } }] });
          const afterXy = state.pointSelections.length;
          const outputVisible = elements.pointSelectionOutputSection.hidden === false;
          const outputHtml = elements.pointSelectionOutput.innerHTML;
          const markerCount = elements.trajectoryXY.data.filter((trace) => trace.meta && trace.meta.pointSelectionMarker).length;
          const firstMarker = elements.trajectoryXY.data.find((trace) => trace.meta && trace.meta.pointSelectionMarker);
          const firstHitTarget = elements.trajectoryXY.data.find((trace) => trace.meta && trace.meta.pointSelectionHitTarget);
          const hitTargetCount = elements.trajectoryXY.data.filter((trace) => trace.meta && trace.meta.pointSelectionHitTarget).length;
          const firstMarkerColor = firstMarker.marker.color;
          const firstMarkerLineWidth = firstMarker.marker.line.width;
          const firstHitTargetSize = firstHitTarget?.marker?.size || 0;
          const firstHitTargetOpacity = firstHitTarget?.marker?.opacity ?? null;
          const firstHitTargetHoverInfo = firstHitTarget?.hoverinfo || null;
          const firstHitTargetHoverTemplate = firstHitTarget?.hovertemplate || null;
          const mutationCallsAfterFirstSelection = plotlyCallStats.addTraceCalls + plotlyCallStats.deleteTraceCalls;
          handlePlotPointHover("trajectoryXY", { points: [{ x: 10, y: 20, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [188.5], xaxis: "x", yaxis: "y" } }] });
          handlePlotPointHover("trajectoryXY", { points: [{ x: 10, y: 20, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [188.5], xaxis: "x", yaxis: "y" } }] });
          const mutationCallsAfterRepeatedSameHover = plotlyCallStats.addTraceCalls + plotlyCallStats.deleteTraceCalls;
          handlePlotPointClick("trajectoryXY", { points: [{ x: 10, y: 20, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [188.5], xaxis: "x", yaxis: "y" } }] });
          const afterClickingSameUnderlyingPoint = state.pointSelections.length;
          state.focusedPointSelectionId = null;
          handlePlotPointHover("trajectoryXY", { points: [{ x: 10, y: 20, pointNumber: 0, data: firstHitTarget }] });
          const focusedByHitTarget = state.focusedPointSelectionId;
          handlePointSelectionKeydown({ key: "Delete", target: { tagName: "INPUT", type: "checkbox" }, preventDefault() {} });
          const afterDelete = state.pointSelections.length;
          state.activePointSelectionChartId = "trajectoryXY";
          handlePlotPointClick("trajectoryXY", { points: [{ x: 11, y: 21, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [189.5], xaxis: "x", yaxis: "y" } }] });
          const reusedMarker = elements.trajectoryXY.data.find((trace) => trace.meta && trace.meta.pointSelectionMarker);
          const reusedMarkerColor = reusedMarker.marker.color;
          const reusedMarkerLineWidth = reusedMarker.marker.line.width;
          handlePlotPointClick("trajectoryXY", { points: [{ x: 12, y: 22, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [190.5], xaxis: "x", yaxis: "y" } }] });
          const twoPointsBeforeSecondDelete = state.pointSelections.length;
          state.activePointSelectionChartId = null;
          state.focusedPointSelectionId = null;
          handlePlotPointHover("trajectoryXY", { points: [{ x: 12, y: 22, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [190.5], xaxis: "x", yaxis: "y" } }] });
          const focusedBeforeSecondDelete = state.focusedPointSelectionId;
          handlePointSelectionKeydown({ key: "Delete", target: { tagName: "INPUT", type: "text" }, preventDefault() {} });
          const afterTextInputDeleteAttempt = state.pointSelections.length;
          handlePointSelectionKeydown({ key: "Delete", target: { tagName: "DIV" }, preventDefault() {} });
          const afterSecondDelete = state.pointSelections.length;
          state.activePointSelectionChartId = "trajectoryXY";
          [
            [13, 23, 191.5],
            [14, 24, 192.5],
            [15, 25, 193.5],
          ].forEach(([x, y, timestamp]) => {
            handlePlotPointClick("trajectoryXY", { points: [{ x, y, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [timestamp], xaxis: "x", yaxis: "y" } }] });
          });
          const multiPointsBeforeBulkDelete = state.pointSelections.length;
          state.activePointSelectionChartId = null;
          [
            [13, 23, 191.5],
            [14, 24, 192.5],
            [11, 21, 189.5],
            [15, 25, 193.5],
          ].forEach(([x, y, timestamp]) => {
            handlePlotPointHover("trajectoryXY", { points: [{ x, y, curveNumber: 1, pointNumber: 0, data: { name: "vloc", customdata: [timestamp], xaxis: "x", yaxis: "y" } }] });
            handlePointSelectionKeydown({ key: "Delete", target: { tagName: "DIV" }, preventDefault() {} });
          });
          const afterBulkDelete = state.pointSelections.length;
          state.activePointSelectionChartId = "trajectoryXY";
          for (let index = 0; index < 11; index += 1) {
            handlePlotPointClick("trajectoryXY", { points: [{
              x: 100 + index,
              y: 200 + index,
              curveNumber: 1,
              pointNumber: 0,
              data: { name: "vloc", customdata: [300 + index], xaxis: "x", yaxis: "y" },
            }] });
          }
          const blackOneSelection = state.pointSelections.find((selection) => selection.markerColor === "#000000" && selection.markerText === "1");
          const blackOneMarker = elements.trajectoryXY.data.find((trace) => trace.meta && trace.meta.selectionId === blackOneSelection.id && trace.meta.pointSelectionMarker);
          state.activePointSelectionChartId = null;
          state.focusedPointSelectionId = null;
          handlePlotPointHover("trajectoryXY", { points: [
            { x: -1, y: -1, curveNumber: 0, pointNumber: 0, data: { name: "unrelated", customdata: [999], xaxis: "x", yaxis: "y" } },
            { x: blackOneSelection.x, y: blackOneSelection.y, pointNumber: 0, data: blackOneMarker },
          ] });
          const focusedBlackOne = state.focusedPointSelectionId;
          handlePointSelectionKeydown({ key: "Delete", target: { tagName: "DIV" }, preventDefault() {} });
          const afterBlackOneDelete = state.pointSelections.length;
          clearAllPointSelections();
          process.stdout.write(JSON.stringify({
            pickableHas3d: PICKABLE_VLOC_CHART_IDS.includes("trajectory3d"),
            pickableHasXy: PICKABLE_VLOC_CHART_IDS.includes("trajectoryXY"),
            after3d,
            afterXy,
            outputVisible,
            outputHtml,
            markerCount,
            hitTargetCount,
            firstMarkerColor,
            reusedMarkerColor,
            firstMarkerLineWidth,
            reusedMarkerLineWidth,
            firstHitTargetSize,
            firstHitTargetOpacity,
            firstHitTargetHoverInfo,
            firstHitTargetHoverTemplate,
            afterClickingSameUnderlyingPoint,
            focusedByHitTarget,
            mutationCallsAfterFirstSelection,
            mutationCallsAfterRepeatedSameHover,
            afterDelete,
            twoPointsBeforeSecondDelete,
            focusedBeforeSecondDelete,
            afterTextInputDeleteAttempt,
            afterSecondDelete,
            multiPointsBeforeBulkDelete,
            afterBulkDelete,
            blackOneSelectionId: blackOneSelection.id,
            focusedBlackOne,
            afterBlackOneDelete,
            afterClear: state.pointSelections.length,
            outputHiddenAfterClear: elements.pointSelectionOutputSection.hidden,
          }));
        `, context);
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["pickableHas3d"] is False
    assert payload["pickableHasXy"] is True
    assert payload["after3d"] == 0
    assert payload["afterXy"] == 1
    assert payload["outputVisible"] is True
    assert "俯视 NE 轨迹" in payload["outputHtml"]
    assert "vloc" in payload["outputHtml"]
    assert "188.500" in payload["outputHtml"]
    assert payload["markerCount"] == 1
    assert payload["hitTargetCount"] == 1
    assert payload["firstHitTargetSize"] > 10
    assert payload["firstHitTargetOpacity"] < 0.1
    assert payload["firstHitTargetHoverInfo"] == "none"
    assert payload["firstHitTargetHoverTemplate"] is None
    assert payload["reusedMarkerColor"] == payload["firstMarkerColor"]
    assert payload["firstMarkerLineWidth"] == 0
    assert payload["reusedMarkerLineWidth"] == 0
    assert payload["afterClickingSameUnderlyingPoint"] == 1
    assert payload["focusedByHitTarget"]
    assert payload["mutationCallsAfterRepeatedSameHover"] == payload["mutationCallsAfterFirstSelection"]
    assert payload["afterDelete"] == 0
    assert payload["twoPointsBeforeSecondDelete"] == 2
    assert payload["focusedBeforeSecondDelete"]
    assert payload["afterTextInputDeleteAttempt"] == 2
    assert payload["afterSecondDelete"] == 1
    assert payload["multiPointsBeforeBulkDelete"] == 4
    assert payload["afterBulkDelete"] == 0
    assert payload["focusedBlackOne"] == payload["blackOneSelectionId"]
    assert payload["afterBlackOneDelete"] == 10
    assert payload["afterClear"] == 0
    assert payload["outputHiddenAfterClear"] is True


def test_static_vo_point_selection_excludes_3d_and_records_points():
    html = Path("static_web/index.html").read_text()
    assert 'id="pointSelectionOutputSection" data-entry-show="vloc,vo"' in html

    source = Path("static_web/app.js").read_text()
    assert 'PICKABLE_VO_CHART_IDS = VO_VISIBLE_CHART_IDS.filter((id) => id !== "trajectory3d")' in source

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          dataset: {},
          data: [],
          classList: { add() {}, remove() {}, toggle() {} },
          addEventListener() {},
          appendChild() {},
          querySelector() { return null; },
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vo"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          voChartDirectorySection: makeElement(),
          voChartList: makeElement(),
          voChartSelectAll: makeElement(),
          voChartClear: makeElement(),
          pointSelectionOutputSection: makeElement(),
          pointSelectionOutput: makeElement(),
          clearAllPointSelections: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          metrics: makeElement(),
          summaryKicker: makeElement(),
          summaryTitle: makeElement(),
          visualKicker: makeElement(),
          visualTitle: makeElement(),
        };
        [
          "trajectory3d", "trajectoryXY", "errorDistance",
          "heightComparison", "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus", "voStatus",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
          "rpeTranslationTime", "rpeRotationTime",
        ].forEach((id) => { elements[id] = makeElement(); });
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return makeElement(); },
          addEventListener() {},
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: {
            addTraces(id, traces) { elements[id].data.push(...traces); },
            deleteTraces(id, indices) { elements[id].data = elements[id].data.filter((_trace, index) => !indices.includes(index)); },
            purge(id) { elements[id].data = []; },
          },
          elements,
          process,
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(`${code}
          state.report = { inputs: { entry_mode: "vo" } };
          state.activePointSelectionChartId = "trajectory3d";
          handlePlotPointClick("trajectory3d", { points: [{ x: 1, y: 2, curveNumber: 0, pointNumber: 0, data: { name: "VO aligned", customdata: [188.5] } }] });
          const after3d = state.pointSelections.length;
          state.activePointSelectionChartId = "positionCompareComposite";
          handlePlotPointClick("positionCompareComposite", { points: [{ x: 188.5, y: 20, curveNumber: 1, pointNumber: 0, data: { name: "X VO aligned", customdata: [188.5], xaxis: "x", yaxis: "y" } }] });
          const afterPosition = state.pointSelections.length;
          const outputVisible = elements.pointSelectionOutputSection.hidden === false;
          const outputHtml = elements.pointSelectionOutput.innerHTML;
          const markerCount = elements.positionCompareComposite.data.filter((trace) => trace.meta && trace.meta.pointSelectionMarker).length;
          state.focusedPointSelectionId = null;
          const marker = elements.positionCompareComposite.data.find((trace) => trace.meta && trace.meta.pointSelectionMarker);
          handlePlotPointHover("positionCompareComposite", { points: [{ x: 188.5, y: 20, pointNumber: 0, data: marker }] });
          const focusedBeforeDelete = state.focusedPointSelectionId;
          handlePointSelectionKeydown({ key: "Delete", target: { tagName: "DIV" }, preventDefault() {} });
          const afterDelete = state.pointSelections.length;
          process.stdout.write(JSON.stringify({
            pickableHas3d: PICKABLE_VO_CHART_IDS.includes("trajectory3d"),
            pickableHasPosition: PICKABLE_VO_CHART_IDS.includes("positionCompareComposite"),
            after3d,
            afterPosition,
            outputVisible,
            outputHtml,
            markerCount,
            focusedBeforeDelete,
            afterDelete,
          }));
        `, context);
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert payload["pickableHas3d"] is False
    assert payload["pickableHasPosition"] is True
    assert payload["after3d"] == 0
    assert payload["afterPosition"] == 1
    assert payload["outputVisible"] is True
    assert "位置随时间变化" in payload["outputHtml"]
    assert "X VO aligned" in payload["outputHtml"]
    assert "188.500" in payload["outputHtml"]
    assert payload["markerCount"] == 1
    assert payload["focusedBeforeDelete"]
    assert payload["afterDelete"] == 0


def test_static_vloc_visuals_use_vloc_detail_tables_and_show_status_charts():
    html = Path("static_web/index.html").read_text()
    for chart_id in ["navStatusModes", "navVelocity", "navResetCounts", "vlocStatus", "heightComparison"]:
        assert f'id="{chart_id}"' in html

    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const plots = [];
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() {},
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          metrics: makeElement(),
        };
        [
          "trajectory3d", "trajectoryXY", "errorDistance",
          "positionCompareComposite", "attitudeCompareComposite",
          "positionErrorComposite", "attitudeErrorComposite",
          "rpeTranslationTime", "rpeRotationTime",
          "navStatusModes", "navVelocity", "navResetCounts", "vlocStatus", "heightComparison",
        ].forEach((id) => { elements[id] = makeElement(); });
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot(id, data, layout) { plots.push({ id, data, layout }); }, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);

        context.renderCharts({
          inputs: { entry_mode: "vloc" },
          per_pose: [
            { timestamp: 10, segment_id: 0, distance_m: 0, gt_x_m: 999, gt_y_m: 999, gt_z_m: 999, est_x_aligned_m: 888, est_y_aligned_m: 888, est_z_aligned_m: 888, x_error_m: 777 },
          ],
          segment_errors: [{ length_m: 100, translation_error_percent: { mean: 99 } }],
          speed_bins: [{ speed_bin_mps: "0-5", translation_error_percent: { mean: 99 } }],
          vloc_details: {
            nav_status: [
              { timestamp: 1, flight_mode: 3, navi_mode: 5, rtk_yaw: 1, rtk_alti: 0, position_reset_count: 0, altitude_reset_count: 1, heading_reset_count: 2, vx: 0.1, vy: 0.2, vz: 0.3, velocity_norm: 0.374 },
              { timestamp: 2, flight_mode: 4, navi_mode: 6, rtk_yaw: 1, rtk_alti: 1, position_reset_count: 0, altitude_reset_count: 1, heading_reset_count: 3, vx: 0.4, vy: 0.5, vz: 0.6, velocity_norm: 0.877 },
            ],
            vloc_status: [
              { timestamp: 1, vloc_mode: 2, num_inliers: 40, reset_count: 0 },
              { timestamp: 2, vloc_mode: 3, num_inliers: 41, reset_count: 1 },
            ],
            comparison: [
              { timestamp: 1, segment_id: 0, visual_segment_id: 0, nav_n_m: 10, nav_e_m: 20, nav_d_m: -3, vloc_n_m: 9, vloc_e_m: 21, vloc_d_m: -4, nav_height_m: 7, vloc_height_m: 6, position_error_n_m: 1, position_error_e_m: -1, position_error_d_m: 1, horizontal_position_error_m: 1.414, attitude_error_yaw_deg: 0.5, attitude_error_pitch_deg: -0.2, attitude_error_roll_deg: 0.1, nav_yaw_deg: 4, vloc_yaw_deg: 3.5, nav_pitch_deg: 1, vloc_pitch_deg: 1.2, nav_roll_deg: 2, vloc_roll_deg: 1.9 },
              { timestamp: 2, segment_id: 1, visual_segment_id: 1, nav_n_m: 11, nav_e_m: 22, nav_d_m: -5, vloc_n_m: 10, vloc_e_m: 23, vloc_d_m: -6, nav_height_m: 8, vloc_height_m: 7, position_error_n_m: 1, position_error_e_m: -1, position_error_d_m: 1, horizontal_position_error_m: 1.414, attitude_error_yaw_deg: 0.6, attitude_error_pitch_deg: -0.3, attitude_error_roll_deg: 0.2, nav_yaw_deg: 5, vloc_yaw_deg: 4.4, nav_pitch_deg: 2, vloc_pitch_deg: 2.3, nav_roll_deg: 3, vloc_roll_deg: 2.8 },
            ],
          },
        });

        const byId = Object.fromEntries(plots.map((plot) => [plot.id, plot]));
        process.stdout.write(JSON.stringify({
          plottedIds: plots.map((plot) => plot.id),
          xChartGt: byId.positionCompareComposite.data[0].y,
          xChartEst: byId.positionCompareComposite.data[1].y,
          xError: byId.positionErrorComposite.data[0].y,
          trajectory3dNames: byId.trajectory3d.data.map((trace) => trace.name),
          vlocStartText: byId.trajectory3d.data.find((trace) => trace.name === "vloc start").text,
          vlocEndText: byId.trajectory3d.data.find((trace) => trace.name === "vloc end").text,
          vlocStartMarker: byId.trajectory3d.data.find((trace) => trace.name === "vloc start").marker,
          vlocStartTextFont: byId.trajectory3d.data.find((trace) => trace.name === "vloc start").textfont,
          positionCompareNames: byId.positionCompareComposite.data.map((trace) => trace.name),
          positionCompareThirdRowColors: [
            byId.positionCompareComposite.data[4].line.color,
            byId.positionCompareComposite.data[5].line.color,
          ],
          heightNames: byId.heightComparison.data.map((trace) => trace.name),
          navVelocityNames: byId.navVelocity.data.map((trace) => trace.name),
          vlocStatusNames: byId.vlocStatus.data.map((trace) => trace.name),
          navStatusHidden: elements.navStatusModes.hidden,
        }));
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    payload = json.loads(result.stdout)
    assert "navStatusModes" in payload["plottedIds"]
    assert "navVelocity" in payload["plottedIds"]
    assert "navResetCounts" in payload["plottedIds"]
    assert "vlocStatus" in payload["plottedIds"]
    assert "heightComparison" in payload["plottedIds"]
    assert payload["xChartGt"] == [10, 11]
    assert payload["xChartEst"] == [9, 10]
    assert payload["xError"] == [1, 1]
    assert "nav start" not in payload["trajectory3dNames"]
    assert "nav end" not in payload["trajectory3dNames"]
    assert {"vloc start", "vloc end"}.issubset(set(payload["trajectory3dNames"]))
    assert payload["vlocStartText"] == ["vloc S1", "vloc S2"]
    assert payload["vlocEndText"] == ["vloc E1", "vloc E2"]
    assert payload["vlocStartMarker"]["color"] == "#9333ea"
    assert payload["vlocStartMarker"]["size"] == 5
    assert payload["vlocStartMarker"]["line"]["width"] == 1
    assert payload["vlocStartTextFont"]["size"] == 10
    assert {"N nav", "N vloc", "E nav", "E vloc", "D nav", "D vloc"} == set(payload["positionCompareNames"])
    assert payload["positionCompareThirdRowColors"] == ["#dc2626", "#0891b2"]
    assert payload["heightNames"] == ["nav height", "vloc height"]
    assert "velocity_norm" in payload["navVelocityNames"]
    assert {"vloc_mode", "reset_count", "num_inliers"}.issubset(set(payload["vlocStatusNames"]))
    assert payload["navStatusHidden"] is False


def test_static_vloc_metrics_hide_vo_specific_summary_cards():
    script = textwrap.dedent(
        r"""
        const fs = require("fs");
        const vm = require("vm");
        const makeElement = (value = "") => ({
          value,
          files: [],
          disabled: false,
          hidden: false,
          textContent: "",
          innerHTML: "",
          style: {},
          classList: { add() {}, remove() {} },
          addEventListener() {},
          click() {},
        });
        const elements = {
          runtimeStatus: makeElement(),
          message: makeElement(),
          runButton: makeElement(),
          entryMode: makeElement("vloc"),
          entryModeHint: makeElement(),
          dataDirFiles: makeElement(),
          logDirFiles: makeElement(),
          dataDirButton: makeElement(),
          logDirButton: makeElement(),
          dataDirStatus: makeElement(),
          logDirStatus: makeElement(),
          downloadJson: makeElement(),
          downloadPoseCsv: makeElement(),
          downloadSegmentCsv: makeElement(),
          downloadWorstCsv: makeElement(),
          downloadConfigJson: makeElement(),
          downloadTrajectoryExcel: makeElement(),
          downloadHtml: makeElement(),
          metrics: makeElement(),
        };
        const document = {
          body: { appendChild() {} },
          getElementById(id) { return elements[id] || makeElement(); },
          createElement() { return { ...makeElement(), remove() {} }; },
          querySelectorAll() { return []; },
        };
        const context = {
          console,
          document,
          window: { location: { protocol: "http:" } },
          TextEncoder,
          Uint8Array,
          DataView,
          Blob: function Blob() {},
          URL: { createObjectURL() { return ""; }, revokeObjectURL() {} },
          Plotly: { newPlot() {}, purge() {} },
        };
        context.globalThis = context;
        const code = fs.readFileSync("static_web/app.js", "utf8").replace(/\ninit\(\);\n/, "\n");
        vm.runInNewContext(code, context);
        context.renderMetrics({
          summary: {
            gt_path_length_m: 1000,
            duration_s: 100,
            matched_poses: 200,
            original_matched_poses: 200,
            gt_pose_coverage_ratio: 0.95,
            est_pose_coverage_ratio: 0.95,
            endpoint_error_m: 1,
            endpoint_error_percent_of_path: 0.1,
            est_poses: 210,
            raw_path_scale_ratio_est_over_gt: 0.32,
          },
          ate_position_m: { rmse: 1.5 },
          ate_vertical_m: { rmse: 0.5 },
          rpe_frame_delta: { translation_m: { rmse: 0.3 }, delta_unit: "frames", delta_frames: 1 },
          divergence: { diverged: false },
          association: { mode: "interpolate_gt", max_interpolation_gap_s: 1.0 },
          discontinuities: { all_matches: { break_count: 2 }, selected_segment: { policy: "segments" } },
          alignment: { scale: 1.0 },
          orientation_correction: { selected: "none" },
          inputs: { entry_mode: "vloc" },
        });
        process.stdout.write(elements.metrics.innerHTML);
        """
    )
    result = subprocess.run(["node", "-e", script], check=True, capture_output=True, text=True)
    html = result.stdout
    assert "ATE RMSE" in html
    assert "mean_error_pos_xy" in html
    assert "mean_error_euler" in html
    assert "时间同步" not in html
    assert "Raw 尺度比" not in html
    assert "对齐尺度" not in html
    assert "姿态修正" not in html
    assert "终点漂移" not in html
    assert "发散状态" not in html
