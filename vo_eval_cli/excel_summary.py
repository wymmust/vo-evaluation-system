"""Small Excel summary writer for CLI batch evaluation."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class Thresholds:
    xy_warn: float = 20.0
    z_warn: float = 20.0
    xy_fail: float = 50.0
    z_fail: float = 50.0
    rpe_trans_warn: float = 0.05
    rpe_trans_fail: float = 0.10


def default_output_path(base_dir: Path | None = None) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return (base_dir or Path.cwd()) / f"eval_summary_{ts}.xlsx"


def write_eval_summary(
    rows: list[dict[str, Any]],
    output_path: Path,
    *,
    rpe_labels: list[str],
    thresholds: Thresholds | None = None,
    primary_key: str | None = None,
) -> Path:
    """Write the evo/Matlab-style summary workbook."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel output. "
            "Install vo-evaluation-system requirements first: "
            "python3 -m pip install -r requirements.txt"
        ) from exc

    thresholds = thresholds or Thresholds()
    primary_key = primary_key or _default_primary_key(rows, rpe_labels)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ok_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    warn_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    fail_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")

    wb = Workbook()
    ws = wb.active
    ws.title = "Eval Summary"

    columns = _columns(rows, rpe_labels)
    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    counts = {"ok": 0, "warn": 0, "fail": 0}
    for row_idx, row in enumerate(rows, 2):
        status = _classify(row, primary_key, thresholds)
        counts[status] += 1
        fill = {"ok": ok_fill, "warn": warn_fill, "fail": fail_fill}[status]

        for col_idx, name in enumerate(columns, 1):
            value = row.get(name, "-")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                cell.alignment = Alignment(vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col_idx >= 3 and status != "ok":
                cell.fill = fill

    widths = [8, 80] + [16] * max(0, len(columns) - 2)
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    _write_statistics_sheet(wb.create_sheet("Statistics"), rows, counts, thresholds, primary_key, thin_border)
    wb.save(output_path)
    return output_path


def _columns(rows: list[dict[str, Any]], rpe_labels: list[str]) -> list[str]:
    fixed = ["log_id", "nav_path"]
    primary_rpe = [f"rpe_{label}_trans_rmse" for label in rpe_labels]
    ape = [
        "_frame_count",
        "ape_trans_max",
        "ape_trans_mean",
        "ape_trans_median",
        "ape_trans_min",
        "ape_trans_rmse",
        "ape_trans_sse",
        "ape_trans_std",
        "ate_trans_rmse",
        "ate_trans_mean",
        "ate_trans_median",
        "ate_trans_min",
        "ate_trans_max",
        "ate_trans_sse",
        "ate_trans_std",
        "mean_xy",
        "mean_z",
        "max_xy",
        "max_z",
        "trajectory_length_m",
        "mean_error_euler",
        "max_error_euler",
    ]
    rpe_details: list[str] = []
    for label in rpe_labels:
        prefix = f"rpe_{label}_trans"
        rpe_details.extend([
            f"{prefix}_max",
            f"{prefix}_mean",
            f"{prefix}_median",
            f"{prefix}_min",
            f"{prefix}_sse",
            f"{prefix}_std",
        ])
    extras = [
        "matched_poses",
        "gt_coverage",
        "est_coverage",
        "duration_s",
        "alignment_scale",
        "valid_est_after_segment_filter",
        "dropped_est_invalid_segment",
        "valid_est_after_mode_filter",
        "dropped_est_invalid_mode",
        "status",
        "message",
    ]
    preferred = fixed + primary_rpe + ape + rpe_details + extras

    all_keys: set[str] = set()
    for row in rows:
        all_keys.update(row.keys())

    columns = [name for name in preferred if name in all_keys or name in fixed]
    columns.extend(sorted(key for key in all_keys if key not in columns))
    return columns


def _default_primary_key(rows: list[dict[str, Any]], rpe_labels: list[str]) -> str:
    all_keys = {key for row in rows for key in row.keys()}
    if "rpe_100m_trans_rmse" in all_keys:
        return "rpe_100m_trans_rmse"
    if "mean_xy" in all_keys:
        return "mean_xy"
    for label in rpe_labels:
        key = f"rpe_{label}_trans_rmse"
        if key in all_keys:
            return key
    if "ape_trans_rmse" in all_keys:
        return "ape_trans_rmse"
    if "ate_trans_rmse" in all_keys:
        return "ate_trans_rmse"
    return "ape_trans_rmse"


def _classify(row: dict[str, Any], primary_key: str, thresholds: Thresholds) -> str:
    status = str(row.get("status", "")).upper()
    if status and status != "OK":
        return "fail"

    value = _as_float(row.get(primary_key))
    if value is None:
        return "ok"

    if primary_key.startswith("rpe_") and "_trans_" in primary_key:
        if value > thresholds.rpe_trans_fail:
            return "fail"
        if value > thresholds.rpe_trans_warn:
            return "warn"
        return "ok"

    if primary_key.startswith(("ape_", "ate_")):
        if value > thresholds.xy_fail:
            return "fail"
        if value > thresholds.xy_warn:
            return "warn"
        return "ok"

    max_xy = _as_float(row.get("max_xy")) or 0.0
    max_z = _as_float(row.get("max_z")) or 0.0
    if max_xy > thresholds.xy_fail:
        return "fail"
    if max_xy > thresholds.xy_warn or max_z > thresholds.z_warn:
        return "warn"
    return "ok"


def _as_float(value: Any) -> float | None:
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    return out


def _write_statistics_sheet(ws, rows, counts, thresholds, primary_key, border) -> None:
    from openpyxl.styles import Font

    total = len(rows)
    pass_rate = f"{counts['ok'] / total * 100:.1f}%" if total else "N/A"
    stats = [
        ("Evaluation Summary Statistics", ""),
        ("", ""),
        ("Total Records", total),
        ("Passed (Green)", counts["ok"]),
        ("Warnings (Yellow)", counts["warn"]),
        ("Failed / Abnormal (Red)", counts["fail"]),
        ("Pass Rate", pass_rate),
        ("Primary Key", primary_key),
        ("", ""),
        ("Threshold Settings", ""),
        ("XY Warning Threshold (m)", thresholds.xy_warn),
        ("Z Warning Threshold (m)", thresholds.z_warn),
        ("XY Failure Threshold (m)", thresholds.xy_fail),
        ("RPE Translation Warning", thresholds.rpe_trans_warn),
        ("RPE Translation Failure", thresholds.rpe_trans_fail),
    ]
    for row_idx, (label, value) in enumerate(stats, 1):
        c1 = ws.cell(row=row_idx, column=1, value=label)
        c2 = ws.cell(row=row_idx, column=2, value=value)
        c1.border = border
        c2.border = border
        if row_idx == 1:
            c1.font = Font(bold=True, size=14)
        elif label and row_idx > 2:
            c1.font = Font(bold=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 36
